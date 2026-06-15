# -*- coding: utf-8 -*-
"""
UC4.5 – Báo cáo tiền lương tất cả bác sĩ, lễ tân, kế toán trong một tháng
14 test cases: uc4.5_01 .. uc4.5_14
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

import requests, pymysql
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.cell import MergedCell

BASE  = "http://localhost:5000"
EXCEL = "D:/OneDrive/Documents/Đánh giá và kiểm định chất lượng phần mềm/Testcase/COUR01.LT2.G06.TestCase.xlsx"
SHEET = "UC04-Nhóm chức năng 4"

GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

REPORT_MONTH = "2026-06"
EMPTY_MONTH  = "2020-05"   # chac chan khong co phieu luong

def db_conn():
    return pymysql.connect(host='localhost', user='root', password='123456',
                           database='dental_clinic', charset='utf8mb4')

def login(u, p):
    r = requests.post(f"{BASE}/api/auth/login", json={"username": u, "password": p})
    if r.status_code == 200:
        d = r.json()
        return d.get("token") or d.get("accessToken") or (d.get("data") or {}).get("token", "")
    return None

def update_excel(results: dict):
    print("\n=== Cap nhat Excel UC4.5 ===")
    try:
        wb = load_workbook(EXCEL)
        ws = wb[SHEET]
        tc_rows = {}
        for row in ws.iter_rows(min_col=1, max_col=1):
            cell = row[0]
            if cell.value:
                v = str(cell.value).strip().lower()
                if v in results:
                    tc_rows[v] = cell.row
        for tc_id, rn in tc_rows.items():
            val  = results[tc_id]
            cell = ws.cell(row=rn, column=7)
            if isinstance(cell, MergedCell):
                continue
            cell.value = val
            cell.fill  = (GREEN if val == "Đạt"
                          else RED if "Không đạt" in val
                          else PatternFill())
        wb.save(EXCEL)
        print(f"Da luu – {len(tc_rows)}/{len(results)} TC cap nhat")
        for tc_id, rn in sorted(tc_rows.items()):
            icon = "✓" if results[tc_id] == "Đạt" else ("✗" if "Không đạt" in results[tc_id] else "○")
            print(f"  {icon} row {rn}: {tc_id} = {results[tc_id]}")
    except Exception as e:
        print(f"Loi luu Excel: {e}")

def main():
    results: dict[str, str] = {}

    # ── Login ──
    admin_token = login("testadmin_x1", "Admin@123")
    if not admin_token:
        print("[ERROR] Khong login duoc admin"); return
    A = {"Authorization": f"Bearer {admin_token}"}

    # Ke toan (letan_tc01)
    rr = requests.post(f"{BASE}/api/auth/users/27/reset-password",
                       json={"newPassword": "Test@12345"}, headers=A)
    Acc = None
    if rr.status_code == 200:
        at = login("letan_tc01", "Test@12345")
        if at:
            Acc = {"Authorization": f"Bearer {at}"}
            print("[SETUP] Ke toan OK")

    # Bac si (bacsi2) cho test khong co quyen
    rr2 = requests.post(f"{BASE}/api/auth/users/14/reset-password",
                        json={"newPassword": "Doctor@Test1"}, headers=A)
    Doc = None
    if rr2.status_code == 200:
        dt = login("bacsi2", "Doctor@Test1")
        if dt:
            Doc = {"Authorization": f"Bearer {dt}"}
            print("[SETUP] Bac si OK")

    # Lay du lieu bao cao lan dau lam tham chieu
    ref_r = requests.get(f"{BASE}/api/salary/report/monthly",
                         params={"month": REPORT_MONTH}, headers=A)
    ref   = ref_r.json() if ref_r.status_code == 200 else {}
    ref_rows  = ref.get("rows", [])
    ref_total = ref.get("totalFund", 0)
    print(f"[INFO] Thang {REPORT_MONTH}: {len(ref_rows)} phieu, totalFund={ref_total:,.0f}")
    print()

    # ════════════════════════════════════════════════════════
    # uc4.5_01 – Admin xem bao cao luong thang thanh cong
    # ════════════════════════════════════════════════════════
    r = requests.get(f"{BASE}/api/salary/report/monthly",
                     params={"month": REPORT_MONTH}, headers=A)
    ok = (r.status_code == 200
          and "rows" in r.json()
          and "totalFund" in r.json()
          and "byRole" in r.json())
    results["uc4.5_01"] = ("Đạt" if ok
                            else f"Không đạt (HTTP {r.status_code}: {r.text[:120]})")
    print(f"uc4.5_01: {r.status_code}, rows={len(r.json().get('rows',[]))} → {results['uc4.5_01']}")

    # ════════════════════════════════════════════════════════
    # uc4.5_02 – Ke toan xem bao cao luong thang thanh cong
    # ════════════════════════════════════════════════════════
    if Acc:
        r = requests.get(f"{BASE}/api/salary/report/monthly",
                         params={"month": REPORT_MONTH}, headers=Acc)
        ok = r.status_code == 200 and "rows" in r.json()
        results["uc4.5_02"] = ("Đạt" if ok
                                else f"Không đạt (HTTP {r.status_code}: {r.text[:120]})")
        print(f"uc4.5_02 (ke toan): {r.status_code} → {results['uc4.5_02']}")
    else:
        results["uc4.5_02"] = "Chưa test (không login được kế toán)"
        print(f"uc4.5_02: → {results['uc4.5_02']}")

    # ════════════════════════════════════════════════════════
    # uc4.5_03 – Hien thi dung cac chi so tong quan
    # totalFund = sum netSalary, byRole counts dung
    # ════════════════════════════════════════════════════════
    if ref_rows:
        data       = ref
        total_fund = data.get("totalFund", 0)
        by_role    = data.get("byRole", {})
        rows       = data.get("rows", [])

        # Kiem tra totalFund = sum(netSalary) cua tat ca rows
        calc_fund = sum(row.get("netSalary", 0) for row in rows)
        ok_fund   = abs(total_fund - calc_fund) < 1

        # Kiem tra byRole count dung
        for role_key in ["DOCTOR", "RECEPTIONIST", "ACCOUNTANT"]:
            expected_cnt = sum(1 for r in rows if r.get("role") == role_key)
            actual_cnt   = by_role.get(role_key, {}).get("count", -1)
            if expected_cnt != actual_cnt:
                ok_fund = False
                print(f"  Role {role_key}: expected={expected_cnt}, actual={actual_cnt}")

        results["uc4.5_03"] = ("Đạt" if ok_fund
                                else f"Không đạt (totalFund={total_fund}, calc={calc_fund})")
        role_summary = ", ".join(f"{k}:{v['count']}" for k, v in by_role.items())
        print(f"uc4.5_03: totalFund={total_fund:,.0f}, calc={calc_fund:,.0f}, byRole={role_summary} → {results['uc4.5_03']}")
    else:
        results["uc4.5_03"] = "Chưa test (không có dữ liệu tháng 06/2026)"
        print(f"uc4.5_03: → {results['uc4.5_03']}")

    # ════════════════════════════════════════════════════════
    # uc4.5_04 – Hien thi bang chi tiet luong tung nhan su
    # Kiem tra moi row co cac field: fullName, role, salaryAmount, netSalary, status
    # ════════════════════════════════════════════════════════
    REQUIRED_FIELDS = ["userId", "fullName", "role", "salaryAmount",
                       "allowance", "deduction", "netSalary", "status"]
    ok_04 = len(ref_rows) > 0
    missing = []
    if ok_04:
        first = ref_rows[0]
        for field in REQUIRED_FIELDS:
            if field not in first:
                missing.append(field)
                ok_04 = False
    results["uc4.5_04"] = ("Đạt" if ok_04
                            else f"Không đạt (thiếu fields: {missing})")
    print(f"uc4.5_04: rows={len(ref_rows)}, fields OK={not missing} → {results['uc4.5_04']}")

    # ════════════════════════════════════════════════════════
    # uc4.5_05 – Loc bao cao theo vai tro Bac si
    # API tra ve tat ca, client loc theo role=DOCTOR
    # Kiem tra: rows filtered by DOCTOR khop byRole[DOCTOR].count
    # ════════════════════════════════════════════════════════
    doctor_rows = [r for r in ref_rows if r.get("role") == "DOCTOR"]
    by_role     = ref.get("byRole", {})
    expected_dr = by_role.get("DOCTOR", {}).get("count", -1)
    ok_05 = len(doctor_rows) == expected_dr and len(doctor_rows) > 0
    results["uc4.5_05"] = ("Đạt" if ok_05
                            else f"Không đạt (filter={len(doctor_rows)}, byRole={expected_dr})")
    print(f"uc4.5_05: DOCTOR filter={len(doctor_rows)}, byRole.count={expected_dr} → {results['uc4.5_05']}")

    # ════════════════════════════════════════════════════════
    # uc4.5_06 – Loc theo Le tan hoac Ke toan
    # ════════════════════════════════════════════════════════
    lt_rows  = [r for r in ref_rows if r.get("role") == "RECEPTIONIST"]
    kt_rows  = [r for r in ref_rows if r.get("role") == "ACCOUNTANT"]
    exp_lt   = by_role.get("RECEPTIONIST", {}).get("count", -1)
    exp_kt   = by_role.get("ACCOUNTANT",   {}).get("count", -1)
    ok_06    = (len(lt_rows) == exp_lt) and (len(kt_rows) == exp_kt)
    results["uc4.5_06"] = ("Đạt" if ok_06
                            else f"Không đạt (LT filter={len(lt_rows)}/exp={exp_lt}, KT={len(kt_rows)}/exp={exp_kt})")
    print(f"uc4.5_06: LT={len(lt_rows)}/{exp_lt}, KT={len(kt_rows)}/{exp_kt} → {results['uc4.5_06']}")

    # ════════════════════════════════════════════════════════
    # uc4.5_07 – Tim kiem nhan su trong bao cao
    # Kiem tra fullName co trong rows → tim kiem client-side
    # ════════════════════════════════════════════════════════
    if ref_rows:
        sample_name = ref_rows[0].get("fullName", "")
        keyword     = sample_name[:3].lower() if len(sample_name) >= 3 else sample_name.lower()
        found       = [r for r in ref_rows if keyword in r.get("fullName", "").lower()]
        ok_07       = len(found) > 0 and all("fullName" in r for r in ref_rows)
        results["uc4.5_07"] = ("Đạt" if ok_07
                                else f"Không đạt (keyword='{keyword}', found={len(found)})")
        print(f"uc4.5_07: search '{keyword}' → {len(found)} ket qua → {results['uc4.5_07']}")
    else:
        results["uc4.5_07"] = "Chưa test (không có dữ liệu)"
        print(f"uc4.5_07: → {results['uc4.5_07']}")

    # ════════════════════════════════════════════════════════
    # uc4.5_08 – Xem chi tiet mot phieu luong tu bao cao
    # GET /payslip/data?userId=X&month=YYYY-MM
    # ════════════════════════════════════════════════════════
    if ref_rows:
        target_user = ref_rows[0].get("userId")
        r = requests.get(f"{BASE}/api/salary/payslip/data",
                         params={"userId": target_user, "month": REPORT_MONTH},
                         headers=A)
        ok = r.status_code == 200 and r.json().get("payslipId") is not None
        detail_fields = ["salaryAmount", "netSalary", "role", "payslipId", "status"]
        if ok:
            ok = all(f in r.json() for f in detail_fields)
        results["uc4.5_08"] = ("Đạt" if ok
                                else f"Không đạt (HTTP {r.status_code}: {r.text[:100]})")
        print(f"uc4.5_08: userId={target_user}, payslipId={r.json().get('payslipId') if r.status_code==200 else '?'} → {results['uc4.5_08']}")
    else:
        results["uc4.5_08"] = "Chưa test (không có dữ liệu)"
        print(f"uc4.5_08: → {results['uc4.5_08']}")

    # ════════════════════════════════════════════════════════
    # uc4.5_09 – Bao cao phan biet trang thai phieu (DRAFT/APPROVED/FINALIZED)
    # ════════════════════════════════════════════════════════
    all_statuses = {r.get("status") for r in ref_rows}
    ok_09 = (len(all_statuses) >= 2 and   # co it nhat 2 trang thai khac nhau
             all(s in ["DRAFT","APPROVED","FINALIZED","CANCELLED"] for s in all_statuses))
    results["uc4.5_09"] = ("Đạt" if ok_09
                            else f"Không đạt (statuses={all_statuses})")
    print(f"uc4.5_09: statuses={all_statuses} → {results['uc4.5_09']}")

    # ════════════════════════════════════════════════════════
    # uc4.5_10 – Xuat Excel → Chua test (frontend feature, khong co API endpoint)
    # uc4.5_11 – Xuat PDF  → Chua test
    # ════════════════════════════════════════════════════════
    results["uc4.5_10"] = "Chưa test (xuất file Excel là tính năng frontend, không có API endpoint)"
    results["uc4.5_11"] = "Chưa test (xuất file PDF là tính năng frontend, không có API endpoint)"
    print(f"uc4.5_10: → {results['uc4.5_10']}")
    print(f"uc4.5_11: → {results['uc4.5_11']}")

    # ════════════════════════════════════════════════════════
    # uc4.5_12 – Bao cao rong khi thang chua co phieu luong
    # ════════════════════════════════════════════════════════
    r = requests.get(f"{BASE}/api/salary/report/monthly",
                     params={"month": EMPTY_MONTH}, headers=A)
    ok = r.status_code == 200
    if ok:
        d = r.json()
        ok = (d.get("totalFund", -1) == 0 and len(d.get("rows", [1])) == 0)
    results["uc4.5_12"] = ("Đạt" if ok
                            else f"Không đạt (HTTP {r.status_code}, rows={len(r.json().get('rows',[]))}, fund={r.json().get('totalFund')})")
    print(f"uc4.5_12: thang {EMPTY_MONTH} → rows={len(r.json().get('rows',[]) if r.status_code==200 else [])}, totalFund={r.json().get('totalFund') if r.status_code==200 else '?'} → {results['uc4.5_12']}")

    # ════════════════════════════════════════════════════════
    # uc4.5_13 – Nguoi khong du quyen khong xem duoc (Bac si → 403)
    # ════════════════════════════════════════════════════════
    if Doc:
        r = requests.get(f"{BASE}/api/salary/report/monthly",
                         params={"month": REPORT_MONTH}, headers=Doc)
        ok = r.status_code == 403
        results["uc4.5_13"] = ("Đạt" if ok
                                else f"Không đạt (HTTP {r.status_code}: {r.text[:120]})")
        print(f"uc4.5_13 (bac si): {r.status_code} → {results['uc4.5_13']}")
    else:
        results["uc4.5_13"] = "Chưa test (không login được bác sĩ)"
        print(f"uc4.5_13: → {results['uc4.5_13']}")

    # ════════════════════════════════════════════════════════
    # uc4.5_14 – Loi xuat file → Chua test (khong the gia lap qua API)
    # ════════════════════════════════════════════════════════
    results["uc4.5_14"] = "Chưa test (không thể giả lập lỗi xuất file qua API)"
    print(f"uc4.5_14: → {results['uc4.5_14']}")

    # ── Cap nhat Excel ──
    update_excel(results)

    # ── Tom tat ──
    print("\n═══════════════ KET QUA UC4.5 ═══════════════")
    dat  = sum(1 for v in results.values() if v == "Đạt")
    kdat = sum(1 for v in results.values() if "Không đạt" in v)
    chua = sum(1 for v in results.values() if "Chưa test" in v)
    for k in sorted(results):
        v    = results[k]
        icon = "✓" if v == "Đạt" else ("✗" if "Không đạt" in v else "○")
        print(f"  {icon} {k}: {v}")
    print(f"\nTong: {dat} Dat | {kdat} Khong dat | {chua} Chua test / 14")

if __name__ == "__main__":
    main()
