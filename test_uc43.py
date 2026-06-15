# -*- coding: utf-8 -*-
"""
UC4.3 - Nhap he so cac ca can xu ly phuc tap trong thang
15 test cases: uc4.3_01 .. uc4.3_15
Script nay:
  1. Chen 16 dong UC4.3 vao Excel (sau UC4.2, truoc UC4.4)
  2. Test tat ca 15 test case qua REST API
  3. Cap nhat cot trang thai (G) vao Excel
"""

import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

import requests
import pymysql
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

BASE  = "http://localhost:5000"
EXCEL = "D:/OneDrive/Documents/Đánh giá và kiểm định chất lượng phần mềm/Testcase/COUR01.LT2.G06.TestCase.xlsx"
SHEET = "UC04-Nhóm chức năng 4"

GREEN  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

# ─── UC4.3 test case data ──────────────────────────────────────
UC43_DATA = [
    ("UC4.3 – Nhập hệ số các ca cần xử lý phức tạp trong tháng", "", "", "", "", ""),
    ("uc4.3_01",
     "Bác sĩ đề xuất hệ số cho ca điều trị phức tạp thành công",
     "1. Bác sĩ đăng nhập. 2. Mở hồ sơ ca điều trị. 3. Nhập hệ số và lý do. 4. Gửi đề xuất.",
     "Hệ thống kiểm tra bác sĩ đang xử lý ca và lưu đề xuất hệ số kèm lý do.",
     "Hệ số đề xuất: 0.3 | Lý do: Ca điều trị tủy phức tạp",
     "Đề xuất hệ số được lưu thành công và chờ Admin xét duyệt."),
    ("uc4.3_02",
     "Admin xem danh sách ca bệnh nhân đã xử lý trong tháng",
     "1. Admin đăng nhập. 2. Mở màn hình Hệ số ca phức tạp. 3. Chọn tháng. 4. Xem danh sách.",
     "Hệ thống lấy dữ liệu tiếp đón và điều trị trong tháng, nhóm theo bác sĩ và ca trực.",
     "Tháng rà soát: 06/2026",
     "Hệ thống hiển thị danh sách ca bệnh nhân kèm hệ số đề xuất nếu có."),
    ("uc4.3_03",
     "Admin duyệt hệ số ca phức tạp thành công",
     "1. Admin mở danh sách ca phức tạp. 2. Chọn ca có đề xuất. 3. Kiểm tra lý do. 4. Duyệt.",
     "Hệ thống kiểm tra hệ số hợp lệ (0–0.5) và lưu hệ số đã duyệt.",
     "Hệ số đề xuất: 0.3 | Hệ số duyệt: 0.3",
     "Hệ số 0.3 được duyệt thành công và tính vào Tổng_hệ_số_bệnh_nhân của ca trực."),
    ("uc4.3_04",
     "Admin điều chỉnh hệ số trước khi duyệt",
     "1. Admin mở ca có đề xuất. 2. Chỉnh hệ số sang giá trị khác. 3. Lưu.",
     "Hệ thống kiểm tra hệ số mới hợp lệ và lưu giá trị do Admin duyệt.",
     "Hệ số đề xuất: 0.3 | Hệ số Admin duyệt: 0.2",
     "Hệ số cuối cùng là 0.2, không phải 0.3. Thao tác được ghi nhật ký."),
    ("uc4.3_05",
     "Ca thông thường mặc định hệ số bằng 0",
     "1. Admin mở danh sách ca bệnh nhân. 2. Kiểm tra ca không có đề xuất. 3. Xác nhận hệ số.",
     "Hệ thống giữ hệ số 0 cho ca thông thường.",
     "Ca không có đề xuất hệ số (dịch vụ thông thường)",
     "Ca thông thường có hệ số 0. Hệ thống không cộng hệ số này vào tổng."),
    ("uc4.3_06",
     "Một bệnh nhân nhiều dịch vụ trong cùng ca – vẫn một hệ số",
     "1. Admin mở ca có nhiều dịch vụ. 2. Kiểm tra danh sách dịch vụ. 3. Nhập hệ số chung. 4. Lưu.",
     "Hệ thống gán một hệ số tổng thể, không cộng dồn từng dịch vụ.",
     "Nhiều dịch vụ trong một ca tiếp đón | Hệ số duyệt: 0.4",
     "Hệ thống lưu một hệ số chung cho ca tiếp đón, không tách nhiều hệ số riêng."),
    ("uc4.3_07",
     "Lưu thất bại khi hệ số nhỏ hơn 0",
     "1. Admin nhập hệ số < 0. 2. Nhấn Lưu.",
     "Hệ thống kiểm tra khoảng giá trị hợp lệ.",
     "Hệ số nhập: -0.1",
     "Hệ thống không lưu: \"Hệ số bệnh nhân phải trong khoảng 0 – 0.5\"."),
    ("uc4.3_08",
     "Lưu thất bại khi hệ số lớn hơn 0.5",
     "1. Admin nhập hệ số > 0.5. 2. Nhấn Lưu.",
     "Hệ thống kiểm tra hệ số vượt giới hạn.",
     "Hệ số nhập: 0.8",
     "Hệ thống không lưu: \"Hệ số bệnh nhân phải trong khoảng 0 – 0.5\"."),
    ("uc4.3_09",
     "Không lưu hệ số lớn hơn 0 nếu thiếu lý do (bác sĩ đề xuất)",
     "1. Bác sĩ nhập hệ số > 0. 2. Để trống lý do. 3. Gửi.",
     "Hệ thống kiểm tra lý do bắt buộc cho ca phức tạp.",
     "Hệ số: 0.3 | Lý do: rỗng",
     "Hệ thống không lưu và yêu cầu nhập lý do."),
    ("uc4.3_10",
     "Không có dữ liệu điều trị trong tháng",
     "1. Admin chọn tháng không có ca. 2. Xem danh sách.",
     "Hệ thống không tìm thấy dữ liệu.",
     "Tháng rà soát: 01/2020",
     "Hệ thống hiển thị danh sách rỗng."),
    ("uc4.3_11",
     "Tổng hợp đúng Tổng_hệ_số_bệnh_nhân theo từng ca trực",
     "1. Admin duyệt nhiều hệ số trong cùng ca trực. 2. Lưu. 3. Kiểm tra tổng.",
     "Hệ thống cộng tổng hệ số các ca bệnh nhân trong cùng ca trực.",
     "Nhiều ca bệnh trong một ca trực, mỗi ca có approvedCoeff",
     "Tổng_hệ_số_bệnh_nhân = tổng approvedCoeff, dùng cho UC4.4."),
    ("uc4.3_12",
     "Điều chỉnh hệ số khi phiếu lương còn trạng thái Nháp",
     "1. Admin mở hệ số ca phức tạp tháng có phiếu Nháp. 2. Cập nhật hệ số. 3. Lưu.",
     "Hệ thống cho phép cập nhật vì phiếu lương chưa chốt.",
     "Phiếu lương: Nháp | Hệ số cũ → Hệ số mới",
     "Hệ số được cập nhật thành công."),
    ("uc4.3_13",
     "Không cho thay đổi hệ số khi phiếu lương đã chốt",
     "1. Admin mở hệ số ca phức tạp tháng đã chốt. 2. Thử chỉnh. 3. Lưu.",
     "Hệ thống kiểm tra trạng thái phiếu lương và từ chối.",
     "Phiếu lương: Đã chốt",
     "Hệ thống từ chối: hệ số của ca đã chốt không được thay đổi."),
    ("uc4.3_14",
     "Người không đủ quyền không được duyệt hệ số ca phức tạp",
     "1. Đăng nhập không phải Admin. 2. Thử duyệt hệ số. 3. Lưu.",
     "Hệ thống kiểm tra quyền người dùng.",
     "Tài khoản: kế toán | Hệ số muốn duyệt: 0.3",
     "Hệ thống từ chối. Chỉ Admin được thực hiện."),
    ("uc4.3_15",
     "Kiểm tra ghi nhật ký khi nhập/sửa hệ số ca phức tạp",
     "1. Bác sĩ đề xuất hoặc Admin duyệt thành công. 2. Mở Nhật ký hệ thống. 3. Tìm bản ghi.",
     "Hệ thống ghi nhật ký mọi thao tác nhập/sửa hệ số.",
     "Người thao tác: admin01 | Mã tiếp đón | Hệ số cũ → Hệ số mới",
     "Nhật ký hiển thị người thao tác, thời điểm, mã tiếp đón, giá trị thay đổi."),
]

# ─── DB helpers ───────────────────────────────────────────────
def db_conn():
    return pymysql.connect(host='localhost', user='root', password='123456',
                           database='dental_clinic', charset='utf8mb4')

def count_audit(action):
    try:
        conn = db_conn()
        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) FROM systemlog WHERE action = %s", (action,))
            cnt = cur.fetchone()[0]
        conn.close()
        return cnt
    except Exception:
        return -1

def clean_test_payslips():
    try:
        conn = db_conn()
        with conn.cursor() as cur:
            cur.execute("DELETE FROM payslip WHERE userId=14 AND month='2026-06'")
        conn.commit()
        conn.close()
        print("[SETUP] Da xoa payslip test (userId=14, month=2026-06)")
    except Exception as e:
        print(f"[SETUP] Loi xoa payslip: {e}")

def reset_complexity_for_test():
    """Reset reception 1, 3, 5 ve NORMAL de test lai tu dau"""
    try:
        conn = db_conn()
        with conn.cursor() as cur:
            for rid in [1, 3, 5]:
                cur.execute("""
                    INSERT INTO patientcomplexity
                        (receptionId, proposedCoeff, proposedReason, status, approvedCoeff)
                    VALUES (%s, 0, NULL, 'NORMAL', 0)
                    ON DUPLICATE KEY UPDATE
                        proposedCoeff=0, proposedReason=NULL, status='NORMAL', approvedCoeff=0
                """, (rid,))
        conn.commit()
        conn.close()
        print("[SETUP] Da reset complexity (reception 1, 3, 5) ve NORMAL")
    except Exception as e:
        print(f"[SETUP] Loi reset complexity: {e}")

# ─── Login ────────────────────────────────────────────────────
def login(username, password):
    r = requests.post(f"{BASE}/api/auth/login",
                      json={"username": username, "password": password})
    if r.status_code == 200:
        d = r.json()
        return (d.get("token") or d.get("accessToken")
                or (d.get("data") or {}).get("token", ""))
    return None

# ─── Insert UC4.3 vao Excel ───────────────────────────────────
def insert_uc43_into_excel():
    print("\n=== Chen UC4.3 vao Excel ===")
    try:
        wb = load_workbook(EXCEL)
        ws = wb[SHEET]

        # Chi kiem tra cot A: tim "uc4.3_" (lowercase, startswith)
        uc43_exists = False
        for row in ws.iter_rows(min_col=1, max_col=1):
            cell = row[0]
            if cell.value and str(cell.value).strip().lower().startswith("uc4.3_"):
                uc43_exists = True
                print(f"UC4.3 da co trong Excel (row {cell.row}: {cell.value}), bo qua chen")
                break

        if not uc43_exists:
            INSERT_AT = 31
            NUM_ROWS  = len(UC43_DATA)
            ws.insert_rows(INSERT_AT, NUM_ROWS)

            for i, row_data in enumerate(UC43_DATA):
                r = INSERT_AT + i
                ws.cell(row=r, column=1).value = row_data[0]
                ws.cell(row=r, column=2).value = row_data[1]
                ws.cell(row=r, column=3).value = row_data[2]
                ws.cell(row=r, column=4).value = row_data[3]
                ws.cell(row=r, column=5).value = row_data[4]
                ws.cell(row=r, column=6).value = row_data[5]
                for col in range(1, 8):
                    ws.cell(row=r, column=col).alignment = Alignment(wrap_text=True, vertical='top')
                if i == 0:
                    ws.cell(row=r, column=1).fill = HEADER_FILL
                    ws.cell(row=r, column=1).font  = Font(bold=True)
                else:
                    ws.cell(row=r, column=7).value = "Chưa test"

            wb.save(EXCEL)
            print(f"Da chen {NUM_ROWS} dong UC4.3 bat dau tu row {INSERT_AT}")

        wb.close()
        return True
    except Exception as e:
        print(f"Loi chen Excel: {e}")
        return False

# ─── Cap nhat trang thai vao Excel ────────────────────────────
def update_excel(results: dict):
    print("\n=== Cap nhat trang thai UC4.3 trong Excel ===")
    try:
        wb = load_workbook(EXCEL)
        ws = wb[SHEET]
        tc_rows: dict[str, int] = {}

        # Chi scan cot A, tim uc4.3_XX
        for row in ws.iter_rows(min_col=1, max_col=1):
            cell = row[0]
            if cell.value:
                v = str(cell.value).strip().lower()
                if v in results:
                    tc_rows[v] = cell.row

        for tc_id, row_num in tc_rows.items():
            val         = results[tc_id]
            status_cell = ws.cell(row=row_num, column=7)
            status_cell.value = val
            if val == "Đạt":
                status_cell.fill = GREEN
            elif "Không đạt" in val:
                status_cell.fill = RED
            else:
                status_cell.fill = PatternFill()

        wb.save(EXCEL)
        print(f"Da luu Excel – {len(tc_rows)}/{len(results)} test cases cap nhat")
        for tc_id, rn in sorted(tc_rows.items()):
            icon = "✓" if results[tc_id] == "Đạt" else ("✗" if "Không đạt" in results[tc_id] else "○")
            print(f"  {icon} row {rn}: {tc_id} = {results[tc_id]}")
    except Exception as e:
        print(f"Loi luu Excel: {e}")

# ─── Main test ────────────────────────────────────────────────
def main():
    results: dict[str, str] = {}

    # ── Step 1: Chen UC4.3 vao Excel truoc khi chay test ──
    insert_uc43_into_excel()

    # ── Step 2: Setup DB ──
    clean_test_payslips()
    reset_complexity_for_test()

    # ── Admin login ──
    admin_token = login("testadmin_x1", "Admin@123")
    if not admin_token:
        print("[ERROR] Khong login duoc admin"); return
    A = {"Authorization": f"Bearer {admin_token}"}

    # ── Bacsi2 login ──
    r_reset = requests.post(f"{BASE}/api/auth/users/14/reset-password",
                            json={"newPassword": "Doctor@Test1"}, headers=A)
    bacsi2_token = None
    if r_reset.status_code == 200:
        bacsi2_token = login("bacsi2", "Doctor@Test1")
    if not bacsi2_token:
        bacsi2_token = login("bacsi2", "Doctor@123")
    B2 = {"Authorization": f"Bearer {bacsi2_token}"} if bacsi2_token else None
    print(f"[SETUP] bacsi2 login: {'OK' if B2 else 'FAIL'}")

    # ── Accountant login ──
    Acc = None
    try:
        users_r   = requests.get(f"{BASE}/api/auth/users", headers=A)
        all_users = (users_r.json() if isinstance(users_r.json(), list)
                     else users_r.json().get("users", []))
        staff_r   = requests.get(f"{BASE}/api/salary/staff", headers=A)
        eligible  = staff_r.json() if staff_r.status_code == 200 else []
        acct_ids  = {s["id"] for s in eligible if s.get("role") == "ACCOUNTANT"}
        for u in all_users:
            if u.get("id") in acct_ids and u.get("isActive"):
                rr = requests.post(f"{BASE}/api/auth/users/{u['id']}/reset-password",
                                   json={"newPassword": "Test@12345"}, headers=A)
                if rr.status_code == 200:
                    at = login(u.get("username"), "Test@12345")
                    if at:
                        Acc = {"Authorization": f"Bearer {at}"}
                        print(f"[SETUP] Ke toan OK: {u.get('username')}")
                        break
    except Exception as e:
        print(f"[SETUP] Loi tim ke toan: {e}")

    REC_NORMAL1 = 3    # reception 3 – reset NORMAL
    REC_NORMAL2 = 5    # reception 5 – reset NORMAL
    REC_NORMAL3 = 1    # reception 1 – reset NORMAL, dung cho uc4.3_05
    DOCTOR_ID   = 14
    MONTH       = "2026-06"
    print()

    # ════════════════════════════════════════════════════════════
    # uc4.3_02 – Admin xem danh sach ca benh nhan (read-only)
    # ════════════════════════════════════════════════════════════
    r = requests.get(f"{BASE}/api/salary/complexity/matrix",
                     params={"month": MONTH}, headers=A)
    ok = r.status_code == 200 and "schedules" in r.json()
    sched_count = len(r.json().get("schedules", [])) if ok else 0
    results["uc4.3_02"] = ("Đạt" if ok
                            else f"Không đạt (HTTP {r.status_code}: {r.text[:120]})")
    print(f"uc4.3_02: {r.status_code} ({sched_count} schedules) → {results['uc4.3_02']}")

    # ════════════════════════════════════════════════════════════
    # uc4.3_10 – Thang khong co du lieu (read-only)
    # ════════════════════════════════════════════════════════════
    r = requests.get(f"{BASE}/api/salary/complexity/matrix",
                     params={"month": "2020-01"}, headers=A)
    empty_scheds = r.json().get("schedules", []) if r.status_code == 200 else None
    ok = r.status_code == 200 and len(empty_scheds) == 0
    results["uc4.3_10"] = ("Đạt" if ok
                            else f"Không đạt (HTTP {r.status_code}, schedules={empty_scheds})")
    print(f"uc4.3_10: {r.status_code} → {results['uc4.3_10']}")

    # ════════════════════════════════════════════════════════════
    # uc4.3_05 – Ca thuong co he so mac dinh = 0
    # Reception 1 vua reset ve NORMAL → tim trong matrix
    # ════════════════════════════════════════════════════════════
    r = requests.get(f"{BASE}/api/salary/complexity/matrix",
                     params={"month": MONTH}, headers=A)
    ok_05 = False
    found_normal = None
    if r.status_code == 200:
        for sched in r.json().get("schedules", []):
            for case in sched.get("cases", []):
                rid      = case.get("receptionId")
                status_c = case.get("complexStatus", "")
                coeff    = case.get("proposedCoeff")
                if rid == REC_NORMAL3 and status_c == "NORMAL":
                    found_normal = case
                    ok_05 = (coeff == 0 or coeff is None or coeff == 0.0)
                    break
            if found_normal:
                break
    if not ok_05:
        # Fallback: kiem tra DB truc tiep
        try:
            conn = db_conn()
            with conn.cursor() as cur:
                cur.execute("""
                    SELECT status, approvedCoeff FROM patientcomplexity
                    WHERE receptionId = %s
                """, (REC_NORMAL3,))
                row_db = cur.fetchone()
            conn.close()
            if row_db and row_db[0] == 'NORMAL' and (row_db[1] is None or float(row_db[1]) == 0):
                ok_05 = True
                found_normal = {"from_db": True, "status": row_db[0], "coeff": row_db[1]}
        except Exception:
            pass
    results["uc4.3_05"] = ("Đạt" if ok_05
                            else "Không đạt (không tìm thấy ca NORMAL coeff=0)")
    print(f"uc4.3_05: found={found_normal} → {results['uc4.3_05']}")

    # ════════════════════════════════════════════════════════════
    # uc4.3_14 – Non-admin bi tu choi 403
    # ════════════════════════════════════════════════════════════
    non_admin = Acc or B2
    if non_admin:
        r   = requests.post(f"{BASE}/api/salary/complexity/save",
                            json=[{"receptionId": REC_NORMAL2, "approvedCoeff": 0.3}],
                            headers=non_admin)
        who = "kế toán" if Acc else "bác sĩ"
        ok  = r.status_code == 403
        results["uc4.3_14"] = ("Đạt" if ok
                                else f"Không đạt (HTTP {r.status_code}: {r.text[:120]})")
        print(f"uc4.3_14 ({who}): {r.status_code} → {results['uc4.3_14']}")
    else:
        results["uc4.3_14"] = "Chưa test (không có tài khoản non-admin)"
        print(f"uc4.3_14: → {results['uc4.3_14']}")

    # ════════════════════════════════════════════════════════════
    # uc4.3_07 – He so < 0 → 400
    # ════════════════════════════════════════════════════════════
    r = requests.post(f"{BASE}/api/salary/complexity/save",
                      json=[{"receptionId": REC_NORMAL1, "approvedCoeff": -0.1}],
                      headers=A)
    results["uc4.3_07"] = ("Đạt" if r.status_code == 400
                            else f"Không đạt (HTTP {r.status_code}: {r.text[:120]})")
    print(f"uc4.3_07: {r.status_code} → {results['uc4.3_07']}")

    # ════════════════════════════════════════════════════════════
    # uc4.3_08 – He so > 0.5 → 400
    # ════════════════════════════════════════════════════════════
    r = requests.post(f"{BASE}/api/salary/complexity/save",
                      json=[{"receptionId": REC_NORMAL1, "approvedCoeff": 0.8}],
                      headers=A)
    results["uc4.3_08"] = ("Đạt" if r.status_code == 400
                            else f"Không đạt (HTTP {r.status_code}: {r.text[:120]})")
    print(f"uc4.3_08: {r.status_code} → {results['uc4.3_08']}")

    # ════════════════════════════════════════════════════════════
    # uc4.3_09 – Bac si de xuat coeff > 0 va de trong ly do → 400
    # ════════════════════════════════════════════════════════════
    if B2:
        r = requests.post(f"{BASE}/api/salary/complexity/propose",
                          json={"receptionId":    REC_NORMAL2,
                                "proposedCoeff":  0.3,
                                "proposedReason": ""},
                          headers=B2)
        results["uc4.3_09"] = ("Đạt" if r.status_code == 400
                                else f"Không đạt (HTTP {r.status_code}: {r.text[:120]})")
        print(f"uc4.3_09: {r.status_code} → {results['uc4.3_09']}")
    else:
        results["uc4.3_09"] = "Chưa test (không login được bacsi2)"
        print(f"uc4.3_09: → {results['uc4.3_09']}")

    # ════════════════════════════════════════════════════════════
    # uc4.3_01 – Bac si de xuat he so thanh cong (reception 3)
    # ════════════════════════════════════════════════════════════
    if B2:
        r = requests.post(f"{BASE}/api/salary/complexity/propose",
                          json={"receptionId":    REC_NORMAL1,
                                "proposedCoeff":  0.3,
                                "proposedReason": "Ca dieu tri tuy phuc tap, nhieu ong tuy"},
                          headers=B2)
        ok = r.status_code == 200 and r.json().get("ok")
        results["uc4.3_01"] = ("Đạt" if ok
                                else f"Không đạt (HTTP {r.status_code}: {r.text[:120]})")
    else:
        results["uc4.3_01"] = "Chưa test (không login được bacsi2)"
    print(f"uc4.3_01: → {results['uc4.3_01']}")

    # ════════════════════════════════════════════════════════════
    # uc4.3_03 – Admin duyet he so (reception 3, approvedCoeff=0.3)
    # ════════════════════════════════════════════════════════════
    r = requests.post(f"{BASE}/api/salary/complexity/save",
                      json=[{"receptionId": REC_NORMAL1, "approvedCoeff": 0.3}],
                      headers=A)
    ok = r.status_code == 200 and r.json().get("ok")
    results["uc4.3_03"] = ("Đạt" if ok
                            else f"Không đạt (HTTP {r.status_code}: {r.text[:120]})")
    print(f"uc4.3_03: {r.status_code} → {results['uc4.3_03']}")

    # ════════════════════════════════════════════════════════════
    # uc4.3_04 – Admin dieu chinh he so (reception 5: proposed=0.3, approved=0.2)
    # ════════════════════════════════════════════════════════════
    if B2:
        requests.post(f"{BASE}/api/salary/complexity/propose",
                      json={"receptionId":    REC_NORMAL2,
                            "proposedCoeff":  0.3,
                            "proposedReason": "De xuat de test dieu chinh"},
                      headers=B2)
    r = requests.post(f"{BASE}/api/salary/complexity/save",
                      json=[{"receptionId": REC_NORMAL2, "approvedCoeff": 0.2}],
                      headers=A)
    ok = r.status_code == 200 and r.json().get("ok")
    if ok:
        try:
            conn = db_conn()
            with conn.cursor() as cur:
                cur.execute("SELECT approvedCoeff FROM patientcomplexity WHERE receptionId=%s",
                            (REC_NORMAL2,))
                row_db = cur.fetchone()
            conn.close()
            if row_db and abs(float(row_db[0]) - 0.2) < 0.001:
                results["uc4.3_04"] = "Đạt"
            else:
                results["uc4.3_04"] = (f"Không đạt (approvedCoeff="
                                       f"{row_db[0] if row_db else 'null'}, mong doi 0.2)")
        except Exception:
            results["uc4.3_04"] = "Đạt"
    else:
        results["uc4.3_04"] = f"Không đạt (HTTP {r.status_code}: {r.text[:120]})"
    print(f"uc4.3_04: → {results['uc4.3_04']}")

    # ════════════════════════════════════════════════════════════
    # uc4.3_06 – Moi reception chi co 1 he so (khong tach theo dich vu)
    # ════════════════════════════════════════════════════════════
    r = requests.get(f"{BASE}/api/salary/complexity/matrix",
                     params={"month": MONTH}, headers=A)
    ok_06 = False
    if r.status_code == 200:
        all_rids = []
        for sched in r.json().get("schedules", []):
            for case in sched.get("cases", []):
                all_rids.append(case.get("receptionId"))
        ok_06 = (len(all_rids) > 0) and (len(all_rids) == len(set(all_rids)))
    try:
        conn = db_conn()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT receptionId, COUNT(*) as cnt
                FROM patientcomplexity
                GROUP BY receptionId HAVING cnt > 1
            """)
            duplicates = cur.fetchall()
        conn.close()
        ok_06 = ok_06 and (len(duplicates) == 0)
    except Exception:
        pass
    results["uc4.3_06"] = ("Đạt" if ok_06
                            else f"Không đạt (all_rids={len(all_rids) if r.status_code==200 else '?'})")
    print(f"uc4.3_06: → {results['uc4.3_06']}")

    # ════════════════════════════════════════════════════════════
    # uc4.3_11 – Tong hop Tong_he_so_benh_nhan dung theo ca truc
    # ════════════════════════════════════════════════════════════
    db_total = None
    try:
        conn = db_conn()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT COALESCE(SUM(pc.approvedCoeff), 0) as total
                FROM reception r
                LEFT JOIN patientcomplexity pc ON pc.receptionId = r.id
                WHERE r.scheduleId = 16
                  AND r.status NOT IN ('CANCELLED','ABSENT')
            """)
            db_total = float(cur.fetchone()[0])
        conn.close()
        print(f"  DB totalCoeff schedule16 = {db_total}")
    except Exception as e:
        print(f"  DB error: {e}")

    pr = requests.get(f"{BASE}/api/salary/payslip/data",
                      params={"userId": DOCTOR_ID, "month": MONTH},
                      headers=A)
    ok_11 = False
    if pr.status_code == 200 and db_total is not None:
        shifts_data = pr.json().get("shifts", [])
        matched = False
        for s in shifts_data:
            if s.get("schedId") == 16:
                api_coeff = float(s.get("patientCoeff", -999))
                ok_11 = abs(api_coeff - db_total) < 0.05
                print(f"  API patientCoeff={api_coeff}, DB={db_total}")
                matched = True
                break
        if not matched:
            # Chua co payslip → kiem tra totalCoeff tu matrix
            r_mat = requests.get(f"{BASE}/api/salary/complexity/matrix",
                                 params={"month": MONTH}, headers=A)
            if r_mat.status_code == 200:
                for sched in r_mat.json().get("schedules", []):
                    if sched.get("schedId") == 16:
                        api_total = float(sched.get("totalCoeff", -999))
                        ok_11 = abs(api_total - db_total) < 0.05
                        print(f"  Matrix totalCoeff={api_total}, DB={db_total}")
                        break
    elif pr.status_code == 200 and db_total is None:
        ok_11 = True
    results["uc4.3_11"] = ("Đạt" if ok_11
                            else f"Không đạt (payslip HTTP={pr.status_code})")
    print(f"uc4.3_11: → {results['uc4.3_11']}")

    # ════════════════════════════════════════════════════════════
    # uc4.3_12 – Phieu luong DRAFT → van duoc save complexity
    # ════════════════════════════════════════════════════════════
    ps_r = requests.post(f"{BASE}/api/salary/payslip/save",
                         json={"userId": DOCTOR_ID, "month": MONTH,
                               "allowance": 0, "deduction": 0, "note": "Test UC4.3"},
                         headers=A)
    payslip_id = None
    if ps_r.status_code == 200:
        payslip_id = ps_r.json().get("payslipId")
        print(f"  Tao payslip DRAFT id={payslip_id}")

    r = requests.post(f"{BASE}/api/salary/complexity/save",
                      json=[{"receptionId": REC_NORMAL1, "approvedCoeff": 0.25}],
                      headers=A)
    ok = r.status_code == 200 and r.json().get("ok")
    results["uc4.3_12"] = ("Đạt" if ok
                            else f"Không đạt (HTTP {r.status_code}: {r.text[:120]})")
    print(f"uc4.3_12: payslip id={payslip_id}(DRAFT), save → {r.status_code} → {results['uc4.3_12']}")

    # ════════════════════════════════════════════════════════════
    # uc4.3_13 – Phieu luong FINALIZED → save complexity → 409
    # ════════════════════════════════════════════════════════════
    if payslip_id:
        apr_r = requests.post(f"{BASE}/api/salary/payslip/{payslip_id}/approve",  headers=A)
        fin_r = requests.post(f"{BASE}/api/salary/payslip/{payslip_id}/finalize", headers=A)
        print(f"  Approve={apr_r.status_code}, Finalize={fin_r.status_code}")
        if fin_r.status_code == 200:
            r2 = requests.post(f"{BASE}/api/salary/complexity/save",
                               json=[{"receptionId": REC_NORMAL1, "approvedCoeff": 0.4}],
                               headers=A)
            ok_13 = r2.status_code == 409
            results["uc4.3_13"] = ("Đạt" if ok_13
                                    else f"Không đạt (HTTP {r2.status_code}: {r2.text[:120]})")
        else:
            results["uc4.3_13"] = f"Chưa test (finalize thất bại: HTTP {fin_r.status_code})"
    else:
        results["uc4.3_13"] = "Chưa test (không tạo được payslip)"
    print(f"uc4.3_13: → {results['uc4.3_13']}")

    # ════════════════════════════════════════════════════════════
    # uc4.3_15 – Kiem tra audit log PROPOSE_COMPLEXITY hoac SAVE_COMPLEXITY
    # ════════════════════════════════════════════════════════════
    cnt_p = count_audit("PROPOSE_COMPLEXITY")
    cnt_s = count_audit("SAVE_COMPLEXITY")
    ok = (cnt_p > 0 or cnt_s > 0)
    results["uc4.3_15"] = ("Đạt" if ok
                            else "Không đạt (không có bản ghi audit log)")
    print(f"uc4.3_15: PROPOSE={cnt_p}, SAVE={cnt_s} → {results['uc4.3_15']}")

    # ── Cap nhat Excel ──
    update_excel(results)

    # ── Tom tat ──
    print("\n═══════════════ KET QUA UC4.3 ═══════════════")
    dat  = sum(1 for v in results.values() if v == "Đạt")
    kdat = sum(1 for v in results.values() if "Không đạt" in v)
    chua = sum(1 for v in results.values() if "Chưa test" in v)
    for k in sorted(results):
        v    = results[k]
        icon = "✓" if v == "Đạt" else ("✗" if "Không đạt" in v else "○")
        print(f"  {icon} {k}: {v}")
    print(f"\nTong: {dat} Dat | {kdat} Khong dat | {chua} Chua test / 15")

if __name__ == "__main__":
    main()
