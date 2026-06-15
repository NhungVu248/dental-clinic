# -*- coding: utf-8 -*-
"""
UC4.1 - Thiet lap muc tien co ban cho mot gio va luong co dinh theo thang
14 test cases: uc4.1_01 .. uc4.1_14
"""

import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

import requests
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

BASE  = "http://localhost:5000"
EXCEL = r"D:\OneDrive\Documents\Đánh giá và kiểm định chất lượng phần mềm\Testcase\COUR01.LT2.G06.TestCase.xlsx"
SHEET = "UC04-Nhóm chức năng 4"

GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# Test dates far in future to avoid conflicting with real data
HR_DATE1_START  = "2099-01-01"
HR_DATE1_END    = "2099-06-30"
HR_DATE2_START  = "2099-07-01"
HR_DATE2_END    = "2099-12-31"
FS_DATE_START   = "2099-01-01"
FS_DATE_END     = "2099-06-30"
FS_DATE2_START  = "2099-07-01"   # for accountant (non-overlapping with receptionist's 2099-01-01~06-30)

# ─────────────────────── DB cleanup ──────────────────────────
def clean_test_data():
    try:
        import pymysql
        conn = pymysql.connect(
            host='localhost', user='root', password='123456', db='dental_clinic', charset='utf8mb4')
        with conn.cursor() as cur:
            cur.execute("DELETE FROM hourlyrate WHERE YEAR(startDate) = 2099")
            cur.execute("DELETE FROM fixedsalary WHERE YEAR(startDate) = 2099")
        conn.commit()
        conn.close()
        print("[SETUP] Đã xóa dữ liệu test cũ (startDate năm 2099)")
    except ImportError:
        print("[SETUP] pymysql chưa cài – bỏ qua cleanup DB")
    except Exception as e:
        print(f"[SETUP] DB cleanup lỗi: {e} – tiếp tục")

# ─────────────────────── Login helper ────────────────────────
def login(username, password):
    r = requests.post(f"{BASE}/api/auth/login",
                      json={"username": username, "password": password})
    if r.status_code == 200:
        d = r.json()
        return (d.get("token")
                or d.get("accessToken")
                or (d.get("data") or {}).get("token", ""))
    return None

# ─────────────────────── Main ────────────────────────────────
def main():
    results: dict[str, str] = {}

    # ── Admin login ──
    admin_token = login("testadmin_x1", "Admin@123")
    if not admin_token:
        print("[ERROR] Không login được admin")
        sys.exit(1)
    A = {"Authorization": f"Bearer {admin_token}"}

    # ── Doctor login ──
    doctor_token = login("nguyenvand", "Doctor@123")
    D = {"Authorization": f"Bearer {doctor_token}"} if doctor_token else None

    # ── Find accountant user & login ──
    users_r = requests.get(f"{BASE}/api/auth/users", headers=A)
    all_users = users_r.json() if isinstance(users_r.json(), list) \
                else users_r.json().get("users", [])

    accountant_user = None
    for u in all_users:
        roles = [r2.get("name") for r2 in u.get("roles", [])]
        if "ACCOUNTANT" in roles and u.get("isActive"):
            accountant_user = u
            break

    Acc = None  # accountant headers
    if accountant_user:
        uid = accountant_user["id"]
        rr = requests.post(f"{BASE}/api/auth/users/{uid}/reset-password",
                           json={"newPassword": "Accountant@123"}, headers=A)
        if rr.status_code == 200:
            at = login(accountant_user.get("username"), "Accountant@123")
            if at:
                Acc = {"Authorization": f"Bearer {at}"}
                print(f"[SETUP] Kế toán login OK: {accountant_user.get('username')}")
    if not Acc:
        print("[SETUP] Không tìm được tài khoản kế toán để test")

    # ── Get eligible staff (RECEPTIONIST, ACCOUNTANT) ──
    staff_r = requests.get(f"{BASE}/api/salary/staff", headers=A)
    eligible = staff_r.json() if staff_r.status_code == 200 else []

    receptionist_id = None
    accountant_id   = None
    for s in eligible:
        if s.get("role") == "RECEPTIONIST" and receptionist_id is None:
            receptionist_id = s["id"]
        if s.get("role") == "ACCOUNTANT" and accountant_id is None:
            accountant_id = s["id"]

    # Doctor ID: nguyenvand (id=8 from known data)
    doctor_id = 8

    print(f"[SETUP] receptionist_id={receptionist_id}  accountant_id={accountant_id}  doctor_id={doctor_id}")

    # ── Clean old test data ──
    clean_test_data()

    # ══════════════════════════════════════════════════════════
    # uc4.1_01 – Admin thiết lập đơn giá giờ thành công
    # ══════════════════════════════════════════════════════════
    r = requests.post(f"{BASE}/api/salary/hourly-rates",
                      json={"amount": 200000,
                            "startDate": HR_DATE1_START,
                            "endDate":   HR_DATE1_END},
                      headers=A)
    ok = r.status_code in (200, 201) and r.json().get("ok")
    results["uc4.1_01"] = "Đạt" if ok \
        else f"Không đạt (HTTP {r.status_code}: {r.text[:120]})"
    print(f"uc4.1_01: {r.status_code} → {results['uc4.1_01']}")

    # ══════════════════════════════════════════════════════════
    # uc4.1_02 – Admin thiết lập lương cố định cho lễ tân
    # ══════════════════════════════════════════════════════════
    if receptionist_id:
        r = requests.post(f"{BASE}/api/salary/fixed-salaries",
                          json={"userId":    receptionist_id,
                                "amount":    8_000_000,
                                "startDate": FS_DATE_START,
                                "endDate":   FS_DATE_END},
                          headers=A)
        ok = r.status_code in (200, 201) and r.json().get("ok")
        results["uc4.1_02"] = "Đạt" if ok \
            else f"Không đạt (HTTP {r.status_code}: {r.text[:120]})"
    else:
        results["uc4.1_02"] = "Chưa test (không có lễ tân)"
    print(f"uc4.1_02: → {results['uc4.1_02']}")

    # ══════════════════════════════════════════════════════════
    # uc4.1_03 – Admin thiết lập lương cố định cho kế toán
    # ══════════════════════════════════════════════════════════
    if accountant_id:
        r = requests.post(f"{BASE}/api/salary/fixed-salaries",
                          json={"userId":    accountant_id,
                                "amount":    10_000_000,
                                "startDate": FS_DATE_START,
                                "endDate":   FS_DATE_END},
                          headers=A)
        ok = r.status_code in (200, 201) and r.json().get("ok")
        results["uc4.1_03"] = "Đạt" if ok \
            else f"Không đạt (HTTP {r.status_code}: {r.text[:120]})"
    else:
        results["uc4.1_03"] = "Chưa test (không có kế toán)"
    print(f"uc4.1_03: → {results['uc4.1_03']}")

    # ══════════════════════════════════════════════════════════
    # uc4.1_04 – Thất bại khi amount <= 0 (hourly rate)
    # ══════════════════════════════════════════════════════════
    r = requests.post(f"{BASE}/api/salary/hourly-rates",
                      json={"amount": 0, "startDate": "2099-07-01"},
                      headers=A)
    ok = r.status_code == 400
    results["uc4.1_04"] = "Đạt" if ok \
        else f"Không đạt (HTTP {r.status_code}: {r.text[:120]})"
    print(f"uc4.1_04: {r.status_code} → {results['uc4.1_04']}")

    # Also test negative amount
    r2 = requests.post(f"{BASE}/api/salary/hourly-rates",
                       json={"amount": -100, "startDate": "2099-07-01"},
                       headers=A)
    if r2.status_code != 400:
        results["uc4.1_04"] += f" [amount=-100 → HTTP {r2.status_code}]"

    # ══════════════════════════════════════════════════════════
    # uc4.1_05 – Thất bại khi lương cố định < 0
    # ══════════════════════════════════════════════════════════
    if receptionist_id:
        r = requests.post(f"{BASE}/api/salary/fixed-salaries",
                          json={"userId":    receptionist_id,
                                "amount":    -1,
                                "startDate": "2099-07-01"},
                          headers=A)
        ok = r.status_code == 400
        results["uc4.1_05"] = "Đạt" if ok \
            else f"Không đạt (HTTP {r.status_code}: {r.text[:120]})"
    else:
        results["uc4.1_05"] = "Chưa test (không có lễ tân)"
    print(f"uc4.1_05: → {results['uc4.1_05']}")

    # ══════════════════════════════════════════════════════════
    # uc4.1_06 – Thất bại khi khoảng thời gian trùng lặp (hourly rate)
    # Tạo rate với startDate=2099-03-01 → trùng với 2099-01-01~06-30
    # ══════════════════════════════════════════════════════════
    r = requests.post(f"{BASE}/api/salary/hourly-rates",
                      json={"amount": 150_000, "startDate": "2099-03-01"},
                      headers=A)
    ok = r.status_code == 409
    results["uc4.1_06"] = "Đạt" if ok \
        else f"Không đạt (HTTP {r.status_code}: {r.text[:120]})"
    print(f"uc4.1_06: {r.status_code} → {results['uc4.1_06']}")

    # ══════════════════════════════════════════════════════════
    # uc4.1_07 – Thất bại khi ngày kết thúc không hợp lệ (endDate <= startDate)
    # ══════════════════════════════════════════════════════════
    r = requests.post(f"{BASE}/api/salary/hourly-rates",
                      json={"amount":    100_000,
                            "startDate": "2099-08-01",
                            "endDate":   "2099-01-01"},   # endDate BEFORE startDate
                      headers=A)
    ok = r.status_code == 400
    results["uc4.1_07"] = "Đạt" if ok \
        else f"Không đạt (HTTP {r.status_code}: {r.text[:120]})"
    print(f"uc4.1_07: {r.status_code} → {results['uc4.1_07']}")

    # Also test endDate == startDate
    r2 = requests.post(f"{BASE}/api/salary/hourly-rates",
                       json={"amount": 100_000, "startDate": "2099-08-01", "endDate": "2099-08-01"},
                       headers=A)
    if r2.status_code != 400:
        results["uc4.1_07"] += f" [equal dates → HTTP {r2.status_code}]"

    # ══════════════════════════════════════════════════════════
    # uc4.1_08 – Không thể thiết lập lương cố định cho bác sĩ
    # (bác sĩ tính lương theo ca giờ)
    # ══════════════════════════════════════════════════════════
    r = requests.post(f"{BASE}/api/salary/fixed-salaries",
                      json={"userId":    doctor_id,
                            "amount":    5_000_000,
                            "startDate": "2099-09-01"},
                      headers=A)
    ok = r.status_code == 400
    results["uc4.1_08"] = "Đạt" if ok \
        else f"Không đạt (HTTP {r.status_code}: {r.text[:120]})"
    print(f"uc4.1_08: {r.status_code} → {results['uc4.1_08']}")

    # ══════════════════════════════════════════════════════════
    # uc4.1_09 – Kế toán chỉ xem được, không sửa được cấu hình
    # GET /api/salary/hourly-rates → 200 (readOnly = ADMIN|ACCOUNTANT)
    # POST /api/salary/hourly-rates → 403 (admin only)
    # ══════════════════════════════════════════════════════════
    if Acc:
        r_get  = requests.get(f"{BASE}/api/salary/hourly-rates", headers=Acc)
        r_post = requests.post(f"{BASE}/api/salary/hourly-rates",
                               json={"amount": 999, "startDate": "2099-10-01"},
                               headers=Acc)
        ok = r_get.status_code == 200 and r_post.status_code == 403
        results["uc4.1_09"] = "Đạt" if ok \
            else f"Không đạt (GET={r_get.status_code}, POST={r_post.status_code})"
    else:
        results["uc4.1_09"] = "Chưa test (không có tài khoản kế toán)"
    print(f"uc4.1_09: → {results['uc4.1_09']}")

    # ══════════════════════════════════════════════════════════
    # uc4.1_10 – Người dùng không phải admin không thể cấu hình lương
    # Dùng token bác sĩ → POST /api/salary/hourly-rates → 403
    # ══════════════════════════════════════════════════════════
    if D:
        r = requests.post(f"{BASE}/api/salary/hourly-rates",
                          json={"amount": 999, "startDate": "2099-11-01"},
                          headers=D)
        ok = r.status_code == 403
        results["uc4.1_10"] = "Đạt" if ok \
            else f"Không đạt (HTTP {r.status_code}: {r.text[:120]})"
        print(f"uc4.1_10: {r.status_code} → {results['uc4.1_10']}")
    else:
        results["uc4.1_10"] = "Chưa test (không login được bác sĩ)"
        print(f"uc4.1_10: → {results['uc4.1_10']}")

    # ══════════════════════════════════════════════════════════
    # uc4.1_11 – Xem lịch sử cấu hình mức tiền cơ bản
    # GET /api/salary/hourly-rates → 200, trả về mảng
    # ══════════════════════════════════════════════════════════
    r = requests.get(f"{BASE}/api/salary/hourly-rates", headers=A)
    ok = r.status_code == 200 and isinstance(r.json(), list)
    count_hr = len(r.json()) if ok else 0
    results["uc4.1_11"] = "Đạt" if ok \
        else f"Không đạt (HTTP {r.status_code})"
    print(f"uc4.1_11: {r.status_code} ({count_hr} bản ghi) → {results['uc4.1_11']}")

    # Cũng xem fixed-salaries
    r_fs = requests.get(f"{BASE}/api/salary/fixed-salaries", headers=A)
    if r_fs.status_code != 200:
        results["uc4.1_11"] += f" [fixed-salaries={r_fs.status_code}]"

    # ══════════════════════════════════════════════════════════
    # uc4.1_12 – Lịch sử cấu hình không bị xóa khi cập nhật
    # Thêm bản ghi mới (2099-07-01 ~ 2099-12-31) và kiểm tra
    # cả 2 bản ghi năm 2099 đều hiển thị trong GET
    # ══════════════════════════════════════════════════════════
    r2 = requests.post(f"{BASE}/api/salary/hourly-rates",
                       json={"amount":    250_000,
                             "startDate": HR_DATE2_START,
                             "endDate":   HR_DATE2_END},
                       headers=A)
    r_list = requests.get(f"{BASE}/api/salary/hourly-rates", headers=A)
    if r_list.status_code == 200:
        rates_list = r_list.json()
        count_2099 = sum(1 for x in rates_list
                         if "2099" in str(x.get("startDate", "")))
        ok = count_2099 >= 2
        results["uc4.1_12"] = "Đạt" if ok \
            else f"Không đạt (chỉ có {count_2099}/2 bản ghi 2099 trong lịch sử)"
    else:
        results["uc4.1_12"] = f"Không đạt (GET={r_list.status_code})"
    print(f"uc4.1_12: → {results['uc4.1_12']}")

    # ══════════════════════════════════════════════════════════
    # uc4.1_13 – Thất bại khi DB lỗi
    # Không thể mô phỏng lỗi DB qua API → đánh dấu Chưa test
    # ══════════════════════════════════════════════════════════
    results["uc4.1_13"] = "Chưa test"
    print("uc4.1_13: Không thể test lỗi DB qua REST API")

    # ══════════════════════════════════════════════════════════
    # uc4.1_14 – Ghi audit log khi cấu hình thay đổi
    # Kiểm tra bảng systemlog có bản ghi action='CREATE_HOURLY_RATE'
    # ══════════════════════════════════════════════════════════
    try:
        import pymysql
        conn = pymysql.connect(
            host='localhost', user='root', password='123456',
            db='dental_clinic', charset='utf8mb4')
        with conn.cursor() as cur:
            cur.execute(
                "SELECT COUNT(*) FROM systemlog WHERE action = 'CREATE_HOURLY_RATE'")
            cnt = cur.fetchone()[0]
        conn.close()
        ok = cnt > 0
        results["uc4.1_14"] = "Đạt" if ok \
            else "Không đạt (không có bản ghi audit log CREATE_HOURLY_RATE)"
    except ImportError:
        results["uc4.1_14"] = "Chưa test (pymysql chưa cài)"
    except Exception as e:
        results["uc4.1_14"] = f"Chưa test (lỗi DB: {e})"
    print(f"uc4.1_14: → {results['uc4.1_14']}")

    # ══════════════════════════════════════════════════════════
    # Cập nhật Excel
    # ══════════════════════════════════════════════════════════
    print("\n=== Cập nhật Excel ===")
    try:
        wb = load_workbook(EXCEL)
        ws = wb[SHEET]

        # Find rows by scanning col A for test case IDs like "uc4.1_01"
        tc_rows: dict[str, int] = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    val = cell.value.strip().lower()
                    if val in results:
                        tc_rows[val] = cell.row
                        break

        # Write results to column G (index 7)
        for tc_id, row_num in tc_rows.items():
            result_val = results[tc_id]
            status_cell = ws.cell(row=row_num, column=7)
            status_cell.value = result_val
            if result_val == "Đạt":
                status_cell.fill = GREEN
            elif "Không đạt" in result_val:
                status_cell.fill = RED

        wb.save(EXCEL)
        print(f"Đã lưu Excel – cập nhật {len(tc_rows)}/{len(results)} test cases")
    except Exception as e:
        print(f"Lỗi lưu Excel: {e}")

    # ══════════════════════════════════════════════════════════
    print("\n═══════════════ KẾT QUẢ UC4.1 ═══════════════")
    dat  = sum(1 for v in results.values() if v == "Đạt")
    kdat = sum(1 for v in results.values() if "Không đạt" in v)
    chua = sum(1 for v in results.values() if "Chưa test" in v)
    for k, v in results.items():
        icon = "✓" if v == "Đạt" else ("✗" if "Không đạt" in v else "○")
        print(f"  {icon} {k}: {v}")
    print(f"\nTổng: {dat} Đạt | {kdat} Không đạt | {chua} Chưa test / 14")

if __name__ == "__main__":
    main()
