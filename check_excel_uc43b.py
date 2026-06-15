import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from openpyxl import load_workbook

EXCEL = "D:/OneDrive/Documents/Đánh giá và kiểm định chất lượng phần mềm/Testcase/COUR01.LT2.G06.TestCase.xlsx"
SHEET = "UC04-Nhóm chức năng 4"

wb = load_workbook(EXCEL)
ws = wb[SHEET]
# Find any cell containing "uc4.3"
print("Cells containing 'uc4.3':")
for row in ws.iter_rows():
    for cell in row:
        if cell.value and "uc4.3" in str(cell.value).lower():
            print(f"  Row {cell.row}, Col {cell.column}: {repr(str(cell.value)[:100])}")
