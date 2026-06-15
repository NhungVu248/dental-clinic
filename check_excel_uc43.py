import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from openpyxl import load_workbook

EXCEL = "D:/OneDrive/Documents/Đánh giá và kiểm định chất lượng phần mềm/Testcase/COUR01.LT2.G06.TestCase.xlsx"
SHEET = "UC04-Nhóm chức năng 4"

wb = load_workbook(EXCEL)
ws = wb[SHEET]
print(f"Max row: {ws.max_row}")
print("\nRows 28-50 col A and G:")
for rn in range(28, 55):
    a = ws.cell(row=rn, column=1).value
    g = ws.cell(row=rn, column=7).value
    if a or g:
        print(f"  Row {rn}: A={repr(str(a)[:60])} | G={repr(str(g)[:30])}")
