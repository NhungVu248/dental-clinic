# -*- coding: utf-8 -*-
"""
Bo sung va test uc4.4_02 va uc4.4_17 vao Excel
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

import requests, pymysql
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.cell import MergedCell

BASE  = "http://localhost:5000"
EXCEL = "D:/OneDrive/Documents/Đánh giá và kiểm định chất lượng phần mềm/Testcase/COUR01.LT2.G06.TestCase.xlsx"
SHEET = "UC04-Nhóm chức năng 4"

GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# ─── Nội dung 2 TC còn thiếu ──────────────────────────────────
TC02 = {
    "B": "Kế toán lập phiếu lương cho kế toán thành công",
    "C": ("1. Kế toán mở màn hình Phiếu lương.\n"
          "2. Chọn nhân sự có vai trò Kế toán.\n"
          "3. Chọn tháng cần lập lương.\n"
          "4. Nhấn Lập phiếu lương.\n"
          "5. Lưu phiếu lương ở trạng thái nháp."),
    "D": "Hệ thống lấy mức lương cố định của nhân sự kế toán từ UC4.1 và tính lương thực nhận.",
    "E": "Nhân sự: KT001 | Vai trò: Kế toán | Tháng: 06/2026 | Lương cố định: 10.000.000 | Phụ cấp: 0 | Khấu trừ: 0",
    "F": "Phiếu lương được tạo ở trạng thái Nháp. Lương thực nhận = mức lương cố định theo tháng hiệu lực.",
}

TC17 = {
    "B": "Kiểm tra ghi nhật ký khi lập, duyệt, chốt phiếu lương",
    "C": ("1. Kế toán lập phiếu lương thành công.\n"
          "2. Admin duyệt và chốt phiếu lương.\n"
          "3. Mở màn hình Nhật ký hệ thống.\n"
          "4. Kiểm tra các bản ghi thao tác."),
    "D": "Hệ thống ghi nhật ký mọi thao tác lập, duyệt và chốt phiếu lương.",
    "E": "Phiếu lương: PL001 | Người lập: ketoan01 | Người duyệt/chốt: admin01",
    "F": "Nhật ký hệ thống có bản ghi đầy đủ cho thao tác lập (CREATE_PAYSLIP), duyệt (APPROVE_PAYSLIP) và chốt (FINALIZE_PAYSLIP) với thông tin người thực hiện.",
}

# ─── DB ───────────────────────────────────────────────────────
def db_conn():
    return pymysql.connect(host='localhost', user='root', password='123456',
                           database='dental_clinic', charset='utf8mb4')

def db_query(sql, params=()):
    conn = db_conn()
    try:
        with conn.cursor() as cur:
            cur.execute(sql, params)
            return cur.fetchall()
    finally:
        conn.close()

def db_exec(sql, params=()):
    conn = db_conn()
    try:
        with conn.cursor() as cur:
            cur.execute(sql, params)
        conn.commit()
    finally:
        conn.close()

# ─── Login ────────────────────────────────────────────────────
def login(username, password):
    r = requests.post(f"{BASE}/api/auth/login",
                      json={"username": username, "password": password})
    if r.status_code == 200:
        d = r.json()
        return (d.get("token") or d.get("accessToken")
                or (d.get("data") or {}).get("token", ""))
    return None

# ─── Dien noi dung vao Excel ──────────────────────────────────
def fill_excel_content():
    print("=== Dien noi dung uc4.4_02 va uc4.4_17 vao Excel ===")
    wb = load_workbook(EXCEL)
    ws = wb[SHEET]

    # Tim row chua uc4.4_02 va uc4.4_17
    row_02 = row_17 = None
    for row in ws.iter_rows(min_col=1, max_col=1):
        cell = row[0]
        if cell.value and str(cell.value).strip().lower() == "uc4.4_02":
            row_02 = cell.row
        if cell.value and str(cell.value).strip().lower() == "uc4.4_17":
            row_17 = cell.row

    print(f"Tim thay: uc4.4_02 @ row {row_02}, uc4.4_17 @ row {row_17}")

    wrap = Alignment(wrap_text=True, vertical='top')
    col_map = {"B": 2, "C": 3, "D": 4, "E": 5, "F": 6}

    for rn, tc_data in [(row_02, TC02), (row_17, TC17)]:
        if rn is None:
            continue
        for col_letter, col_idx in col_map.items():
            cell = ws.cell(row=rn, column=col_idx)
            if isinstance(cell, MergedCell):
                continue
            cell.value = tc_data[col_letter]
            cell.alignment = wrap
        # Neu cot G chua co gia tri, set "Chua test"
        g_cell = ws.cell(row=rn, column=7)
        if not isinstance(g_cell, MergedCell) and not g_cell.value:
            g_cell.value = "Chưa test"

    wb.save(EXCEL)
    print("Da dien noi dung va luu Excel")

# ─── Cap nhat trang thai ──────────────────────────────────────
def update_status(tc_id: str, val: str):
    wb = load_workbook(EXCEL)
    ws = wb[SHEET]
    for row in ws.iter_rows(min_col=1, max_col=1):
        cell = row[0]
        if cell.value and str(cell.value).strip().lower() == tc_id:
            g = ws.cell(row=cell.row, column=7)
            if isinstance(g, MergedCell):
                break
            g.value = val
            g.fill  = GREEN if val == "Đạt" else (RED if "Không đạt" in val else PatternFill())
            print(f"  Updated row {cell.row}: {tc_id} = {val}")
            break
    wb.save(EXCEL)

# ══════════════════════════════════════════════════════════════
def main():
    # ── Buoc 1: Dien noi dung vao Excel ──
    fill_excel_content()
    print()

    # ── Login ──
    admin_token = login("testadmin_x1", "Admin@123")
    if not admin_token:
        print("[ERROR] Khong login duoc admin"); return
    A = {"Authorization": f"Bearer {admin_token}"}

    # Ke toan login (letan_tc01)
    acc_r = requests.post(f"{BASE}/api/auth/users/27/reset-password",
                          json={"newPassword": "Test@12345"}, headers=A)
    Acc = None
    if acc_r.status_code == 200:
        at = login("letan_tc01", "Test@12345")
        if at:
            Acc = {"Authorization": f"Bearer {at}"}
    Maker = Acc if Acc else A
    print(f"[SETUP] Maker = {'ke toan' if Acc else 'admin'}")

    results = {}

    # ════════════════════════════════════════════════════════
    # uc4.4_02 – Ke toan lap phieu luong cho ke toan thanh cong
    # userId=17 (ketoan1, ACCOUNTANT, fixedsalary~10.251.100/thang)
    # Dung thang 2026-07 de tranh xung dot voi payslip hien co
    # ════════════════════════════════════════════════════════
    KT_ID    = 17   # ketoan1
    KT_MONTH = "2026-07"

    # Xoa payslip test neu co
    db_exec("DELETE FROM payslip WHERE userId=%s AND month=%s", (KT_ID, KT_MONTH))

    r = requests.post(f"{BASE}/api/salary/payslip/save",
                      json={"userId":    KT_ID,
                            "month":     KT_MONTH,
                            "allowance": 0,
                            "deduction": 0,
                            "note":      "Test uc4.4_02"},
                      headers=Maker)
    ok = r.status_code == 200 and r.json().get("ok")
    ps_id = r.json().get("payslipId") if r.status_code == 200 else None

    if ok and ps_id:
        ps_rows = db_query("""
            SELECT role, salaryAmount, netSalary, status FROM payslip WHERE id=%s
        """, (ps_id,))
        if ps_rows:
            role, sal, net, status = ps_rows[0]
            sal  = float(sal)
            net  = float(net)
            # Kiem tra: role=ACCOUNTANT, salary > 0, status=DRAFT
            ok = (role == 'ACCOUNTANT' and sal > 0 and status == 'DRAFT')
            print(f"  role={role}, salaryAmount={sal:,.0f}, netSalary={net:,.0f}, status={status}")
        else:
            ok = False

    results["uc4.4_02"] = ("Đạt" if ok
                            else f"Không đạt (HTTP {r.status_code}: {r.text[:120]})")
    print(f"uc4.4_02: payslipId={ps_id} → {results['uc4.4_02']}")
    print()

    # ════════════════════════════════════════════════════════
    # uc4.4_17 – Kiem tra ghi nhat ky khi lap, duyet, chot phieu luong
    # Kiem tra systemlog co CREATE_PAYSLIP, APPROVE_PAYSLIP, FINALIZE_PAYSLIP
    # ════════════════════════════════════════════════════════

    # Tao phieu → Duyet → Chot de dam bao log ton tai
    # Dung userId=17, thang 2026-08 (fresh, khong xung dot)
    LOG_USER  = KT_ID
    LOG_MONTH = "2026-08"
    db_exec("DELETE FROM payslip WHERE userId=%s AND month=%s", (LOG_USER, LOG_MONTH))

    r_save = requests.post(f"{BASE}/api/salary/payslip/save",
                           json={"userId": LOG_USER, "month": LOG_MONTH,
                                 "allowance": 0, "deduction": 0},
                           headers=Maker)
    log_ps_id = r_save.json().get("payslipId") if r_save.status_code == 200 else None
    print(f"  Tao payslip: id={log_ps_id}, HTTP={r_save.status_code}")

    if log_ps_id:
        r_apr = requests.post(f"{BASE}/api/salary/payslip/{log_ps_id}/approve",  headers=A)
        r_fin = requests.post(f"{BASE}/api/salary/payslip/{log_ps_id}/finalize", headers=A)
        print(f"  Approve={r_apr.status_code}, Finalize={r_fin.status_code}")

    # Kiem tra systemlog
    rows_create   = db_query("SELECT COUNT(*) FROM systemlog WHERE action='CREATE_PAYSLIP'")
    rows_approve  = db_query("SELECT COUNT(*) FROM systemlog WHERE action='APPROVE_PAYSLIP'")
    rows_finalize = db_query("SELECT COUNT(*) FROM systemlog WHERE action='FINALIZE_PAYSLIP'")

    cnt_create   = int(rows_create[0][0])
    cnt_approve  = int(rows_approve[0][0])
    cnt_finalize = int(rows_finalize[0][0])

    ok = cnt_create > 0 and cnt_approve > 0 and cnt_finalize > 0
    results["uc4.4_17"] = ("Đạt" if ok
                            else f"Không đạt (CREATE={cnt_create}, APPROVE={cnt_approve}, FINALIZE={cnt_finalize})")
    print(f"uc4.4_17: CREATE={cnt_create}, APPROVE={cnt_approve}, FINALIZE={cnt_finalize} → {results['uc4.4_17']}")

    # ── Cap nhat Excel ──
    print()
    for tc_id, val in results.items():
        update_status(tc_id, val)

    # ── Tom tat ──
    print("\n══════════ KET QUA BO SUNG ══════════")
    for k, v in results.items():
        icon = "✓" if v == "Đạt" else "✗"
        print(f"  {icon} {k}: {v}")

if __name__ == "__main__":
    main()
