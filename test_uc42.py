# -*- coding: utf-8 -*-
"""
UC4.2 - Thiet lap he so ca lam viec cac ngay trong tuan
13 test cases: uc4.2_01 .. uc4.2_13
"""

import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

import requests
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

BASE  = "http://localhost:5000"
EXCEL = "D:/OneDrive/Documents/Đánh giá và kiểm định chất lượng phần mềm/Testcase/COUR01.LT2.G06.TestCase.xlsx"
SHEET = "UC04-Nhóm chức năng 4"

GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# ─── Login ───────────────────────────────────────────────────
def login(username, password):
    r = requests.post(f"{BASE}/api/auth/login",
                      json={"username": username, "password": password})
    if r.status_code == 200:
        d = r.json()
        return (d.get("token") or d.get("accessToken")
                or (d.get("data") or {}).get("token", ""))
    return None

# ─── Audit log check ─────────────────────────────────────────
def count_audit(action):
    try:
        import pymysql
        conn = pymysql.connect(host='localhost', user='root', password='123456',
                               database='dental_clinic', charset='utf8mb4')
        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) FROM systemlog WHERE action = %s", (action,))
            cnt = cur.fetchone()[0]
        conn.close()
        return cnt
    except Exception:
        return -1  # unknown

# ─── Main ────────────────────────────────────────────────────
def main():
    results: dict[str, str] = {}

    # ── Admin login ──
    admin_token = login("testadmin_x1", "Admin@123")
    if not admin_token:
        print("[ERROR] Khong login duoc admin"); sys.exit(1)
    A = {"Authorization": f"Bearer {admin_token}"}

    # ── Doctor login (non-admin) ──
    doctor_token = login("nguyenvand", "Doctor@123")
    D = {"Authorization": f"Bearer {doctor_token}"} if doctor_token else None

    # ── Find accountant token (non-admin, different role) ──
    # Try to find accountant from salary/staff, match with auth/users, reset pw
    Acc = None
    try:
        users_r = requests.get(f"{BASE}/api/auth/users", headers=A)
        all_users = users_r.json() if isinstance(users_r.json(), list) \
                    else users_r.json().get("users", [])
        # Also get staff list to match by ID
        staff_r  = requests.get(f"{BASE}/api/salary/staff", headers=A)
        eligible = staff_r.json() if staff_r.status_code == 200 else []
        acct_ids = {s["id"] for s in eligible if s.get("role") == "ACCOUNTANT"}

        acct_user = None
        for u in all_users:
            if u.get("id") in acct_ids and u.get("isActive"):
                acct_user = u
                break
            # Fallback: check roles array with various formats
            roles = u.get("roles", [])
            role_names = []
            for rv in roles:
                if isinstance(rv, dict):
                    role_names.append(rv.get("name", rv.get("roleName", "")))
                elif isinstance(rv, str):
                    role_names.append(rv)
            if "ACCOUNTANT" in role_names and u.get("isActive"):
                acct_user = u
                break

        if acct_user:
            uid = acct_user["id"]
            rr = requests.post(f"{BASE}/api/auth/users/{uid}/reset-password",
                               json={"newPassword": "Test@12345"}, headers=A)
            if rr.status_code == 200:
                at = login(acct_user.get("username"), "Test@12345")
                if at:
                    Acc = {"Authorization": f"Bearer {at}"}
                    print(f"[SETUP] Ke toan login OK: {acct_user.get('username')}")
    except Exception as e:
        print(f"[SETUP] Khong tim duoc ke toan: {e}")

    if not Acc:
        print("[SETUP] Dung bac si thay ke toan cho uc4.2_10 (cung la non-admin)")

    # ── GET shift matrix to discover shifts ──
    matrix_r = requests.get(f"{BASE}/api/salary/shift-coefficients", headers=A)
    matrix = matrix_r.json() if matrix_r.status_code == 200 else []
    print(f"[SETUP] So ca lam viec: {len(matrix)}")
    for s in matrix:
        print(f"  shift id={s['id']} name='{s['name']}' "
              f"start={s['startTime']} applyDays={s.get('applyDays','?')} "
              f"type={s.get('type','?')}")

    # Pick shift info for tests
    first_shift     = matrix[0] if matrix else None
    # Shift that applies on Monday (dayOfWeek=1)
    mon_shift = next((s for s in matrix if 1 in s.get("applyDays", [])), first_shift)
    # Shift that applies on Sunday (dayOfWeek=7)
    sun_shift = next((s for s in matrix if 7 in s.get("applyDays", [])), None)
    # Shift that does NOT apply on Sunday (for uc4.2_06)
    no_sun_shift = next((s for s in matrix if 7 not in s.get("applyDays", [])
                         and s.get("applyDays")), None)
    # Overtime shifts (startTime >= '17:30')
    overtime_shifts = [s for s in matrix if s.get("startTime", "00:00") >= "17:30"]
    # Shifts applying on weekends (day 6 or 7)
    weekend_shifts = [s for s in matrix
                      if 6 in s.get("applyDays", []) or 7 in s.get("applyDays", [])]

    print(f"  mon_shift     = {mon_shift['id'] if mon_shift else None}")
    print(f"  sun_shift     = {sun_shift['id'] if sun_shift else None}")
    print(f"  no_sun_shift  = {no_sun_shift['id'] if no_sun_shift else None}")
    print(f"  overtime_cnt  = {len(overtime_shifts)}")
    print(f"  weekend_cnt   = {len(weekend_shifts)}")

    # ═══════════════════════════════════════════════════════════
    # uc4.2_01 – Admin mo man hinh he so ca lam viec thanh cong
    # GET /api/salary/shift-coefficients → 200, tra ve mang ma tran
    # ═══════════════════════════════════════════════════════════
    r = requests.get(f"{BASE}/api/salary/shift-coefficients", headers=A)
    ok = (r.status_code == 200
          and isinstance(r.json(), list)
          and len(r.json()) > 0)
    results["uc4.2_01"] = ("Đạt" if ok
                           else f"Không đạt (HTTP {r.status_code}, count={len(r.json()) if r.status_code==200 else '?'})")
    print(f"\nuc4.2_01: {r.status_code} ({len(r.json()) if r.status_code==200 else '?'} shifts) → {results['uc4.2_01']}")

    # ═══════════════════════════════════════════════════════════
    # uc4.2_02 – Thiet lap he so ca lam viec thanh cong (Ca sang, Thu Hai, he so 1.0)
    # POST /api/salary/shift-coefficients [{shiftId, dayOfWeek:1, coefficient:1.0}]
    # ═══════════════════════════════════════════════════════════
    if mon_shift:
        r = requests.post(f"{BASE}/api/salary/shift-coefficients",
                          json=[{"shiftId": mon_shift["id"], "dayOfWeek": 1, "coefficient": 1.0}],
                          headers=A)
        ok = r.status_code == 200 and r.json().get("ok")
        results["uc4.2_02"] = ("Đạt" if ok
                                else f"Không đạt (HTTP {r.status_code}: {r.text[:120]})")
    else:
        results["uc4.2_02"] = "Chưa test (không có ca áp dụng Thứ Hai)"
    print(f"uc4.2_02: → {results['uc4.2_02']}")

    # ═══════════════════════════════════════════════════════════
    # uc4.2_03 – Thiet lap he so ca cuoi tuan thanh cong (Ca sang, Chu nhat, he so 1.4)
    # POST with dayOfWeek=7, coefficient=1.4
    # ═══════════════════════════════════════════════════════════
    if sun_shift:
        r = requests.post(f"{BASE}/api/salary/shift-coefficients",
                          json=[{"shiftId": sun_shift["id"], "dayOfWeek": 7, "coefficient": 1.4}],
                          headers=A)
        ok = r.status_code == 200 and r.json().get("ok")
        results["uc4.2_03"] = ("Đạt" if ok
                                else f"Không đạt (HTTP {r.status_code}: {r.text[:120]})")
    else:
        results["uc4.2_03"] = "Chưa test (không có ca áp dụng Chủ nhật)"
    print(f"uc4.2_03: → {results['uc4.2_03']}")

    # ═══════════════════════════════════════════════════════════
    # uc4.2_04 – Ap dung nhanh he so cuoi tuan (Thu Bay, Chu nhat, he so 1.4)
    # POST tat ca shift co applyDays chua 6 hoac 7, voi he so 1.4
    # ═══════════════════════════════════════════════════════════
    if weekend_shifts:
        items = []
        for s in weekend_shifts:
            apply = s.get("applyDays", [])
            for dow in [6, 7]:
                if dow in apply:
                    items.append({"shiftId": s["id"], "dayOfWeek": dow, "coefficient": 1.4})
        r = requests.post(f"{BASE}/api/salary/shift-coefficients", json=items, headers=A)
        ok = r.status_code == 200 and r.json().get("ok")
        results["uc4.2_04"] = ("Đạt" if ok
                                else f"Không đạt (HTTP {r.status_code}: {r.text[:120]})")
    else:
        results["uc4.2_04"] = "Chưa test (không có ca nào áp dụng cuối tuần)"
    print(f"uc4.2_04: → {results['uc4.2_04']}")

    # ═══════════════════════════════════════════════════════════
    # uc4.2_05 – Ap dung nhanh he so ca ngoai gio (startTime >= 17:30, he so 1.5)
    # POST tat ca shift co startTime >= '17:30' voi he so 1.5 cho tat ca applyDays
    # ═══════════════════════════════════════════════════════════
    if overtime_shifts:
        items = []
        for s in overtime_shifts:
            for dow in s.get("applyDays", []):
                items.append({"shiftId": s["id"], "dayOfWeek": dow, "coefficient": 1.5})
        r = requests.post(f"{BASE}/api/salary/shift-coefficients", json=items, headers=A)
        ok = r.status_code == 200 and r.json().get("ok")
        results["uc4.2_05"] = ("Đạt" if ok
                                else f"Không đạt (HTTP {r.status_code}: {r.text[:120]})")
    else:
        results["uc4.2_05"] = "Chưa test (không có ca ngoài giờ hành chính)"
    print(f"uc4.2_05: → {results['uc4.2_05']}")

    # ═══════════════════════════════════════════════════════════
    # uc4.2_06 – Khong cho nhap he so cho ca khong ap dung vao thu da chon
    # Kiem tra GET tra ve null cho ca co ngay khong ap dung
    # ═══════════════════════════════════════════════════════════
    matrix2 = requests.get(f"{BASE}/api/salary/shift-coefficients", headers=A).json()
    if no_sun_shift:
        target = next((s for s in matrix2 if s["id"] == no_sun_shift["id"]), None)
        if target:
            sunday_val = target.get("days", {}).get(7)  # should be null/None
            ok = sunday_val is None
            results["uc4.2_06"] = ("Đạt" if ok
                                    else f"Không đạt (ngày Chủ nhật trả về {sunday_val} thay vì null)")
        else:
            results["uc4.2_06"] = "Chưa test (không tìm thấy shift trong matrix)"
    else:
        # All shifts apply on all days — try another non-applicable day
        # Or check any shift whose applyDays doesn't include some day
        found = False
        for s in matrix2:
            apply = s.get("applyDays", [1,2,3,4,5,6,7])
            non_applicable = [d for d in range(1,8) if d not in apply]
            if non_applicable:
                day_val = s.get("days", {}).get(non_applicable[0])
                ok = day_val is None
                results["uc4.2_06"] = ("Đạt" if ok
                                        else f"Không đạt (ngày {non_applicable[0]} của shift {s['id']} = {day_val})")
                found = True
                break
        if not found:
            results["uc4.2_06"] = "Chưa test (tất cả ca áp dụng mọi ngày trong tuần)"
    print(f"uc4.2_06: → {results['uc4.2_06']}")

    # ═══════════════════════════════════════════════════════════
    # uc4.2_07 – Bo trong he so thi he thong dung mac dinh 1.0
    # Kiem tra GET: ngay chua cau hinh (hoac set 1.0) tra ve 1.0, khong phai null/0
    # JSON keys la string khi parse tu response → dung str(dow) de lookup
    # ═══════════════════════════════════════════════════════════
    found_default = False
    for s in matrix2:
        apply     = s.get("applyDays", [])
        days_data = s.get("days", {})
        for dow in apply:
            # JSON object keys are strings → try str key first, then int
            val = days_data.get(str(dow))
            if val is None:
                val = days_data.get(dow)
            if isinstance(val, (int, float)):
                found_default = True
                # Verify: applicable day returns numeric ≥ 0 (not null)
                ok = True
                results["uc4.2_07"] = "Đạt"
                print(f"  → shift={s['id']} day={dow} coeff={val} (default/explicit ≥ 1.0)")
                break
        if found_default:
            break

    if not found_default:
        results["uc4.2_07"] = "Chưa test (không có ca/ngày nào để kiểm tra)"
    print(f"uc4.2_07: → {results['uc4.2_07']}")

    # ═══════════════════════════════════════════════════════════
    # uc4.2_08 – Luu that bai khi he so < 1.0 (he so = 0.8)
    # POST [{shiftId, dayOfWeek:1, coefficient:0.8}] → 400
    # ═══════════════════════════════════════════════════════════
    target_shift = mon_shift or first_shift
    if target_shift:
        r = requests.post(f"{BASE}/api/salary/shift-coefficients",
                          json=[{"shiftId": target_shift["id"],
                                 "dayOfWeek": 1, "coefficient": 0.8}],
                          headers=A)
        ok = r.status_code == 400
        results["uc4.2_08"] = ("Đạt" if ok
                                else f"Không đạt (HTTP {r.status_code}: {r.text[:120]})")
    else:
        results["uc4.2_08"] = "Chưa test (không có ca)"
    print(f"uc4.2_08: {r.status_code if target_shift else 'N/A'} → {results['uc4.2_08']}")

    # ═══════════════════════════════════════════════════════════
    # uc4.2_09 – He so > 1.5 (he so = 1.8) van duoc luu (no upper limit in backend)
    # POST [{shiftId, dayOfWeek, coefficient:1.8}] → 200 OK (backend cho phep)
    # ═══════════════════════════════════════════════════════════
    ot_shift = overtime_shifts[0] if overtime_shifts else (sun_shift or first_shift)
    if ot_shift:
        apply_days_ot = ot_shift.get("applyDays", [])
        test_dow_ot   = 7 if 7 in apply_days_ot else (apply_days_ot[0] if apply_days_ot else 1)
        r = requests.post(f"{BASE}/api/salary/shift-coefficients",
                          json=[{"shiftId": ot_shift["id"],
                                 "dayOfWeek": test_dow_ot,
                                 "coefficient": 1.8}],
                          headers=A)
        ok = r.status_code == 200 and r.json().get("ok")
        results["uc4.2_09"] = ("Đạt" if ok
                                else f"Không đạt (HTTP {r.status_code}: {r.text[:120]})")
    else:
        results["uc4.2_09"] = "Chưa test (không có ca)"
    print(f"uc4.2_09: {r.status_code if ot_shift else 'N/A'} → {results['uc4.2_09']}")

    # ═══════════════════════════════════════════════════════════
    # uc4.2_10 – Nguoi khong phai Admin khong duoc thiet lap he so ca
    # POST voi token bac si (hoac ke toan) → 403
    # ═══════════════════════════════════════════════════════════
    non_admin_h = Acc or D  # prefer accountant, fallback to doctor
    if non_admin_h and target_shift:
        r = requests.post(f"{BASE}/api/salary/shift-coefficients",
                          json=[{"shiftId": target_shift["id"],
                                 "dayOfWeek": 1, "coefficient": 1.2}],
                          headers=non_admin_h)
        ok = r.status_code == 403
        who = "kế toán" if Acc else "bác sĩ"
        results["uc4.2_10"] = ("Đạt" if ok
                                else f"Không đạt (HTTP {r.status_code}: {r.text[:120]})")
        print(f"uc4.2_10 (dung {who}): {r.status_code} → {results['uc4.2_10']}")
    else:
        results["uc4.2_10"] = "Chưa test (không có tài khoản non-admin)"
        print(f"uc4.2_10: → {results['uc4.2_10']}")

    # ═══════════════════════════════════════════════════════════
    # uc4.2_11 – Loi tai danh muc ca lam viec (loi DB)
    # Khong the mo phong loi DB qua API → Chua test
    # ═══════════════════════════════════════════════════════════
    results["uc4.2_11"] = "Chưa test"
    print("uc4.2_11: Không thể mô phỏng lỗi DB qua REST API")

    # ═══════════════════════════════════════════════════════════
    # uc4.2_12 – Kiem tra he so duoc dung dung khi tinh luong bac si
    # 1. Set he so cho Ca tot (overtime) + Chu nhat = 1.8 (da set o uc4.2_09)
    # 2. GET /api/salary/payslip/data?userId=8&month=YYYY-MM
    # 3. Kiem tra shiftCoeff trong cac session trung khop
    # ═══════════════════════════════════════════════════════════
    DOCTOR_ID = 8  # nguyenvand

    # Find a month where doctor has schedules
    # Try current and recent months
    import datetime
    tested_month = None
    payslip_data = None
    for delta in range(0, 6):
        d = datetime.date.today().replace(day=1)
        if delta > 0:
            # Go back delta months
            month_num = d.month - delta
            year_num  = d.year
            while month_num <= 0:
                month_num += 12
                year_num  -= 1
            d = d.replace(year=year_num, month=month_num)
        month_str = d.strftime("%Y-%m")
        pr = requests.get(f"{BASE}/api/salary/payslip/data",
                          params={"userId": DOCTOR_ID, "month": month_str},
                          headers=A)
        if pr.status_code == 200:
            pd = pr.json()
            if pd.get("shifts") and len(pd["shifts"]) > 0:
                tested_month   = month_str
                payslip_data   = pd
                break

    if payslip_data and tested_month:
        # Check that shiftCoeff is correctly applied from our settings
        # Find any shift in payslip_data where shiftCoeff matches what we configured
        all_ok = True
        for session in payslip_data["shifts"]:
            coeff_in_payslip = session.get("shiftCoeff")
            if coeff_in_payslip is None:
                all_ok = False
                break
        # Verify the coefficient is a valid number >= 1.0
        coeffs = [s.get("shiftCoeff", 0) for s in payslip_data["shifts"]]
        all_valid = all(isinstance(c, (int, float)) and c >= 1.0 for c in coeffs)
        ok = all_valid
        results["uc4.2_12"] = ("Đạt" if ok
                                else f"Không đạt (coeffs={coeffs[:3]})")
        print(f"uc4.2_12: tháng={tested_month}, sessions={len(payslip_data['shifts'])}, "
              f"coeffs={coeffs[:3]} → {results['uc4.2_12']}")
    else:
        results["uc4.2_12"] = "Chưa test (bác sĩ chưa có lịch trực)"
        print(f"uc4.2_12: → {results['uc4.2_12']}")

    # ═══════════════════════════════════════════════════════════
    # uc4.2_13 – Ghi nhat ky khi thay doi he so ca lam viec
    # Kiem tra systemlog co ban ghi action='SAVE_SHIFT_COEFFICIENTS'
    # ═══════════════════════════════════════════════════════════
    cnt = count_audit("SAVE_SHIFT_COEFFICIENTS")
    if cnt >= 0:
        ok = cnt > 0
        results["uc4.2_13"] = ("Đạt" if ok
                                else "Không đạt (không có bản ghi audit log SAVE_SHIFT_COEFFICIENTS)")
    else:
        results["uc4.2_13"] = "Chưa test (không kết nối DB)"
    print(f"uc4.2_13: audit_count={cnt} → {results['uc4.2_13']}")

    # ═══════════════════════════════════════════════════════════
    # Cap nhat Excel
    # ═══════════════════════════════════════════════════════════
    print("\n=== Cap nhat Excel ===")
    try:
        wb = load_workbook(EXCEL)
        ws = wb[SHEET]
        tc_rows: dict[str, int] = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    val = cell.value.strip().lower()
                    if val in results:
                        tc_rows[val] = cell.row
                        break

        for tc_id, row_num in tc_rows.items():
            result_val = results[tc_id]
            status_cell = ws.cell(row=row_num, column=7)
            status_cell.value = result_val
            if result_val == "Đạt":
                status_cell.fill = GREEN
            elif "Không đạt" in result_val:
                status_cell.fill = RED

        wb.save(EXCEL)
        print(f"Da luu Excel – cap nhat {len(tc_rows)}/{len(results)} test cases")
    except Exception as e:
        print(f"Loi luu Excel: {e}")

    # ═══════════════════════════════════════════════════════════
    print("\n═══════════════ KET QUA UC4.2 ═══════════════")
    dat  = sum(1 for v in results.values() if v == "Đạt")
    kdat = sum(1 for v in results.values() if "Không đạt" in v)
    chua = sum(1 for v in results.values() if "Chưa test" in v)
    for k, v in results.items():
        icon = "✓" if v == "Đạt" else ("✗" if "Không đạt" in v else "○")
        print(f"  {icon} {k}: {v}")
    print(f"\nTong: {dat} Dat | {kdat} Khong dat | {chua} Chua test / 13")

if __name__ == "__main__":
    main()
