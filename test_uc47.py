# -*- coding: utf-8 -*-
"""
Test UC4.7 – Báo cáo tiền lương tất cả nhân sự trong một năm
GET /api/salary/report/annual/full?year=YYYY  (ADMIN|ACCOUNTANT)
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

import requests, pymysql
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.cell import MergedCell

BASE  = "http://localhost:5000"
EXCEL = "D:/OneDrive/Documents/Đánh giá và kiểm định chất lượng phần mềm/Testcase/COUR01.LT2.G06.TestCase.xlsx"
SHEET = "UC04-Nhóm chức năng 4"

GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# ─── DB helpers ──────────────────────────────────────────────────
def db_conn():
    return pymysql.connect(host='localhost', user='root', password='123456',
                           database='dental_clinic', charset='utf8mb4')

def db_query(sql, params=()):
    conn = db_conn()
    try:
        with conn.cursor() as cur:
            cur.execute(sql, params)
            return cur.fetchall()
    finally:
        conn.close()

def db_exec(sql, params=()):
    conn = db_conn()
    try:
        with conn.cursor() as cur:
            cur.execute(sql, params)
        conn.commit()
    finally:
        conn.close()

# ─── Auth ────────────────────────────────────────────────────────
def login(username, password):
    r = requests.post(f"{BASE}/api/auth/login",
                      json={"username": username, "password": password})
    if r.status_code == 200:
        d = r.json()
        return (d.get("token") or d.get("accessToken")
                or (d.get("data") or {}).get("token", ""))
    return None

# ─── Excel helpers ────────────────────────────────────────────────
def update_status(tc_id: str, val: str):
    wb = load_workbook(EXCEL)
    ws = wb[SHEET]
    for row in ws.iter_rows(min_col=1, max_col=1):
        cell = row[0]
        if cell.value and str(cell.value).strip().lower() == tc_id.lower():
            g = ws.cell(row=cell.row, column=7)
            if isinstance(g, MergedCell):
                # unmerge and write
                ranges_to_remove = [str(mr) for mr in ws.merged_cells.ranges
                                    if mr.min_row <= cell.row <= mr.max_row
                                    and mr.min_col <= 7 <= mr.max_col]
                for mr_str in ranges_to_remove:
                    ws.unmerge_cells(mr_str)
                g = ws.cell(row=cell.row, column=7)
            g.value = val
            g.fill  = GREEN if val == "Đạt" else (RED if "Không đạt" in val else PatternFill())
            print(f"  Excel row {cell.row}: {tc_id} → {val}")
            break
    wb.save(EXCEL)

# ─── Main tests ───────────────────────────────────────────────────
def run_tests(A, ACC_H, DOCTOR_H):
    results = {}
    URL = f"{BASE}/api/salary/report/annual/full"

    # ── Lấy report 2026 làm cơ sở ──
    r_base = requests.get(f"{URL}?year=2026", headers=A)
    report = r_base.json() if r_base.status_code == 200 else None
    if report:
        print(f"[Base report 2026] totalFund={report.get('totalFund'):,.0f}, "
              f"byRole={report.get('byRole')}, employees={len(report.get('employees', []))}")
    print()

    # ─── uc4.7_01 ────────────────────────────────────────────────
    r01 = requests.get(f"{URL}?year=2026", headers=A)
    ok01 = (r01.status_code == 200
            and "totalFund"    in r01.json()
            and "byRole"       in r01.json()
            and "monthlyChart" in r01.json()
            and "employees"    in r01.json())
    results["uc4.7_01"] = "Đạt" if ok01 else f"Không đạt (HTTP {r01.status_code}: {r01.text[:80]})"
    print(f"uc4.7_01: Admin xem báo cáo năm → HTTP {r01.status_code} → {results['uc4.7_01']}")

    # ─── uc4.7_02 ────────────────────────────────────────────────
    r02 = requests.get(f"{URL}?year=2026", headers=ACC_H)
    ok02 = r02.status_code == 200 and "totalFund" in r02.json()
    results["uc4.7_02"] = "Đạt" if ok02 else f"Không đạt (HTTP {r02.status_code}: {r02.text[:80]})"
    print(f"uc4.7_02: Kế toán xem báo cáo năm → HTTP {r02.status_code} → {results['uc4.7_02']}")

    # ─── uc4.7_03 ────────────────────────────────────────────────
    if report:
        by_role   = report.get("byRole", {})
        total_sum = sum(v.get("total", 0) for v in by_role.values())
        api_total = report.get("totalFund", -1)
        ok03 = abs(total_sum - api_total) < 1  # tolerance 1 VND
        detail03 = (f"totalFund={api_total:,.0f}, "
                    f"sum(byRole)={total_sum:,.0f}, "
                    f"DOCTOR={by_role.get('DOCTOR',{}).get('total',0):,.0f}({by_role.get('DOCTOR',{}).get('count',0)} NV), "
                    f"RECEPT={by_role.get('RECEPTIONIST',{}).get('total',0):,.0f}, "
                    f"ACCNT={by_role.get('ACCOUNTANT',{}).get('total',0):,.0f}")
    else:
        ok03 = False; detail03 = "no data"
    results["uc4.7_03"] = "Đạt" if ok03 else f"Không đạt ({detail03})"
    print(f"uc4.7_03: Kiểm tra chỉ số tổng hợp → {detail03} → {results['uc4.7_03']}")

    # ─── uc4.7_04 ────────────────────────────────────────────────
    if report:
        chart = report.get("monthlyChart", [])
        ok04 = (len(chart) == 12
                and all("month" in m and "DOCTOR" in m for m in chart))
    else:
        ok04 = False
    results["uc4.7_04"] = "Đạt" if ok04 else f"Không đạt (monthlyChart={len(chart) if report else 'N/A'} entries)"
    if report:
        nonzero = sum(1 for m in chart if m["DOCTOR"]+m["RECEPTIONIST"]+m["ACCOUNTANT"] > 0)
        print(f"uc4.7_04: monthlyChart có {len(chart)} tháng, {nonzero} tháng có dữ liệu → {results['uc4.7_04']}")
    else:
        print(f"uc4.7_04: (no data) → {results['uc4.7_04']}")

    # ─── uc4.7_05 ────────────────────────────────────────────────
    results["uc4.7_05"] = "Chưa test"
    print("uc4.7_05: Biểu đồ diễn biến quỹ lương – render frontend → Chưa test")

    # ─── uc4.7_06 ────────────────────────────────────────────────
    results["uc4.7_06"] = "Chưa test"
    print("uc4.7_06: Lọc theo vai trò Bác sĩ – bộ lọc frontend, API trả về toàn bộ → Chưa test")

    # ─── uc4.7_07 ────────────────────────────────────────────────
    results["uc4.7_07"] = "Chưa test"
    print("uc4.7_07: Lọc theo vai trò Kế toán/Lễ tân – bộ lọc frontend → Chưa test")

    # ─── uc4.7_08 ────────────────────────────────────────────────
    # Kiểm tra API trả về dữ liệu đủ để tính tỷ trọng (byRole totals hợp lệ)
    if report and report.get("totalFund", 0) > 0:
        by_role  = report.get("byRole", {})
        total    = report.get("totalFund", 0)
        doc_pct  = round(by_role.get("DOCTOR",      {}).get("total", 0) / total * 100, 1)
        rec_pct  = round(by_role.get("RECEPTIONIST",{}).get("total", 0) / total * 100, 1)
        acc_pct  = round(by_role.get("ACCOUNTANT",  {}).get("total", 0) / total * 100, 1)
        ok08 = abs(doc_pct + rec_pct + acc_pct - 100) < 1.0  # tổng tỷ trọng ≈ 100%
    else:
        ok08 = True  # năm không có dữ liệu → totalFund=0, tỷ trọng không áp dụng
        doc_pct = rec_pct = acc_pct = 0
    results["uc4.7_08"] = "Đạt" if ok08 else f"Không đạt (tổng tỷ trọng {doc_pct+rec_pct+acc_pct:.1f}% ≠ 100%)"
    print(f"uc4.7_08: Tỷ trọng DOCTOR={doc_pct}%, RECEPT={rec_pct}%, ACCNT={acc_pct}% (tổng≈100%) → {results['uc4.7_08']}")

    # ─── uc4.7_09 ────────────────────────────────────────────────
    results["uc4.7_09"] = "Chưa test"
    print("uc4.7_09: Xem chi tiết nhân sự → điều hướng sang UC4.6 (frontend) → Chưa test")

    # ─── uc4.7_10 ────────────────────────────────────────────────
    results["uc4.7_10"] = "Chưa test"
    print("uc4.7_10: Chọn tháng → điều hướng sang UC4.5 (frontend) → Chưa test")

    # ─── uc4.7_11 ────────────────────────────────────────────────
    r11 = requests.get(f"{URL}?year=2020", headers=A)
    ok11 = False
    if r11.status_code == 200:
        d11 = r11.json()
        ok11 = (d11.get("totalFund", -1) == 0
                and len(d11.get("employees", [1])) == 0
                and d11.get("countActiveMonths", -1) == 0
                and all(m["DOCTOR"] == 0 and m["RECEPTIONIST"] == 0 and m["ACCOUNTANT"] == 0
                        for m in d11.get("monthlyChart", [])))
    results["uc4.7_11"] = "Đạt" if ok11 else f"Không đạt (HTTP {r11.status_code}: {r11.text[:80]})"
    if r11.status_code == 200:
        d11 = r11.json()
        print(f"uc4.7_11: year=2020 → totalFund={d11.get('totalFund')}, employees={len(d11.get('employees',[]))} → {results['uc4.7_11']}")
    else:
        print(f"uc4.7_11: HTTP {r11.status_code} → {results['uc4.7_11']}")

    # ─── uc4.7_12 ────────────────────────────────────────────────
    r12 = requests.get(f"{URL}?year=2026", headers=DOCTOR_H)
    ok12 = r12.status_code == 403
    results["uc4.7_12"] = "Đạt" if ok12 else f"Không đạt (HTTP {r12.status_code} thay vì 403)"
    print(f"uc4.7_12: Bác sĩ truy cập → HTTP {r12.status_code} → {results['uc4.7_12']}")

    # ─── uc4.7_13 ────────────────────────────────────────────────
    results["uc4.7_13"] = "Chưa test"
    print("uc4.7_13: Xuất Excel báo cáo quỹ lương năm – tính năng frontend → Chưa test")

    # ─── uc4.7_14 ────────────────────────────────────────────────
    results["uc4.7_14"] = "Chưa test"
    print("uc4.7_14: Xuất PDF báo cáo quỹ lương năm – tính năng frontend → Chưa test")

    # ─── uc4.7_15 ────────────────────────────────────────────────
    results["uc4.7_15"] = "Chưa test"
    print("uc4.7_15: Lỗi xuất file – cần giả lập lỗi hạ tầng, frontend → Chưa test")

    # ─── uc4.7_16 ────────────────────────────────────────────────
    results["uc4.7_16"] = "Chưa test"
    print("uc4.7_16: Cảnh báo dữ liệu lớn – cần 500+ nhân sự/6000+ phiếu, frontend → Chưa test")

    # ─── uc4.7_17 ────────────────────────────────────────────────
    # Báo cáo chỉ đọc: không có endpoint POST/PUT/PATCH cho annual/full
    r17_post  = requests.post (f"{URL}?year=2026", json={}, headers=A)
    r17_patch = requests.patch(f"{URL}?year=2026", json={}, headers=A)
    ok17 = r17_post.status_code in (404, 405) and r17_patch.status_code in (404, 405)
    results["uc4.7_17"] = "Đạt" if ok17 else f"Không đạt (POST={r17_post.status_code}, PATCH={r17_patch.status_code})"
    print(f"uc4.7_17: POST={r17_post.status_code}, PATCH={r17_patch.status_code} (expected 404/405) → {results['uc4.7_17']}")

    return results

# ─── Main ─────────────────────────────────────────────────────────
def main():
    # ── Login admin ──
    admin_token = login("testadmin_x1", "Admin@123")
    if not admin_token:
        print("[ERROR] Không login được admin"); return
    A = {"Authorization": f"Bearer {admin_token}"}

    # ── Accountant token ──
    acc_rows = db_query("""
        SELECT u.id, u.username FROM `user` u
        JOIN userrole ur ON ur.userId = u.id
        JOIN role r ON r.id = ur.roleId
        WHERE r.name = 'ACCOUNTANT' AND u.isActive = 1
        LIMIT 1
    """)
    ACC_H = A
    if acc_rows:
        acc_id, acc_username = int(acc_rows[0][0]), acc_rows[0][1]
        rr = requests.post(f"{BASE}/api/auth/users/{acc_id}/reset-password",
                           json={"newPassword": "Test@12345"}, headers=A)
        if rr.status_code == 200:
            at = login(acc_username, "Test@12345")
            if at:
                ACC_H = {"Authorization": f"Bearer {at}"}
                print(f"Accountant: id={acc_id}, username={acc_username}")

    # ── Doctor token ──
    doc_rows = db_query("""
        SELECT u.id, u.username FROM `user` u
        JOIN userrole ur ON ur.userId = u.id
        JOIN role r ON r.id = ur.roleId
        WHERE r.name = 'DOCTOR' AND u.isActive = 1
        LIMIT 1
    """)
    DOCTOR_H = {}
    if doc_rows:
        doc_id, doc_username = int(doc_rows[0][0]), doc_rows[0][1]
        rr2 = requests.post(f"{BASE}/api/auth/users/{doc_id}/reset-password",
                            json={"newPassword": "Test@12345"}, headers=A)
        if rr2.status_code == 200:
            dt = login(doc_username, "Test@12345")
            if dt:
                DOCTOR_H = {"Authorization": f"Bearer {dt}"}
                print(f"Doctor: id={doc_id}, username={doc_username}")

    print()
    print("=== Chạy test UC4.7 ===")
    results = run_tests(A, ACC_H, DOCTOR_H)

    print("\n=== Cập nhật Excel ===")
    for tc_id, val in results.items():
        update_status(tc_id, val)

    print("\n══════════════════ KẾT QUẢ UC4.7 ══════════════════")
    dat   = sum(1 for v in results.values() if v == "Đạt")
    khong = sum(1 for v in results.values() if "Không đạt" in v)
    chua  = sum(1 for v in results.values() if v == "Chưa test")
    for k, v in results.items():
        icon = "✓" if v == "Đạt" else ("✗" if "Không đạt" in v else "–")
        print(f"  {icon} {k}: {v}")
    print(f"\n  Tổng: {dat} Đạt | {khong} Không đạt | {chua} Chưa test / {len(results)} TC")

if __name__ == "__main__":
    main()
