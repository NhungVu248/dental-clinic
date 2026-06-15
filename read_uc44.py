import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from openpyxl import load_workbook

EXCEL = "D:/OneDrive/Documents/Đánh giá và kiểm định chất lượng phần mềm/Testcase/COUR01.LT2.G06.TestCase.xlsx"
SHEET = "UC04-Nhóm chức năng 4"

wb = load_workbook(EXCEL)
ws = wb[SHEET]
print(f"Max row: {ws.max_row}\n")

# Find UC4.4 section
in_uc44 = False
for rn in range(1, ws.max_row + 1):
    a = ws.cell(row=rn, column=1).value
    b = ws.cell(row=rn, column=2).value
    c = ws.cell(row=rn, column=3).value
    d = ws.cell(row=rn, column=4).value
    e = ws.cell(row=rn, column=5).value
    f = ws.cell(row=rn, column=6).value
    g = ws.cell(row=rn, column=7).value
    if a and "uc4.4" in str(a).lower():
        in_uc44 = True
    if in_uc44:
        print(f"Row {rn}:")
        print(f"  A={str(a)[:80]}")
        print(f"  B={str(b)[:80]}")
        print(f"  C={str(c)[:100]}")
        print(f"  D={str(d)[:100]}")
        print(f"  E={str(e)[:100]}")
        print(f"  F={str(f)[:100]}")
        print(f"  G={str(g)}")
        print()
    if in_uc44 and a and "uc4.5" in str(a).lower():
        break
