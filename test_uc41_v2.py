# -*- coding: utf-8 -*-
"""
UC4.1 v2 - Fixed date strategy
- Dung dates 2025 (truoc data thuc 2026) tranh xung dot
- Tim ke toan qua staff list (letan_tc01)
"""

import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

import requests
import pymysql
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

BASE  = "http://localhost:5000"
EXCEL = "D:/OneDrive/Documents/Đánh giá và kiểm định chất lượng phần mềm/Testcase/COUR01.LT2.G06.TestCase.xlsx"
SHEET = "UC04-Nhóm chức năng 4"

GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# Dates an toan: truoc tat ca data thuc (2026-01-01)
# Khong chong lan voi hourlyrate(2026-01-01 ~ 2026-05-31) va (2026-06-01 ~ null)
HR_START1  = "2025-01-01"
HR_END1    = "2025-06-30"
HR_START2  = "2025-07-01"   # dung cho uc4.1_12 (second record)
HR_END2    = "2025-12-31"
FS_START   = "2025-01-01"
FS_END     = "2025-06-30"

# ─── DB helpers ─────────────────────────────────────────────
def db_conn():
    return pymysql.connect(
        host='localhost', user='root', password='123456',
        database='dental_clinic', charset='utf8mb4')

def clean_test_data():
    """Xoa du lieu test nam 2025 (an toan vi data thuc bat dau tu 2026)"""
    try:
        conn = db_conn()
        with conn.cursor() as cur:
            cur.execute("DELETE FROM hourlyrate WHERE YEAR(startDate) = 2025")
            cur.execute("DELETE FROM fixedsalary WHERE YEAR(startDate) = 2025")
            # Xoa luon record 2099 con lai tu test truoc
            cur.execute("DELETE FROM hourlyrate WHERE YEAR(startDate) = 2099")
            cur.execute("DELETE FROM fixedsalary WHERE YEAR(startDate) = 2099")
        conn.commit()
        conn.close()
        print("[SETUP] Da xoa du lieu test cu (nam 2025/2099)")
    except Exception as e:
        print(f"[SETUP] Loi cleanup: {e}")

def count_audit(action):
    try:
        conn = db_conn()
        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) FROM systemlog WHERE action = %s", (action,))
            cnt = cur.fetchone()[0]
        conn.close()
        return cnt
    except Exception:
        return -1

# ─── Login ───────────────────────────────────────────────────
def login(username, password):
    r = requests.post(f"{BASE}/api/auth/login",
                      json={"username": username, "password": password})
    if r.status_code == 200:
        d = r.json()
        return (d.get("token") or d.get("accessToken")
                or (d.get("data") or {}).get("token", ""))
    return None

# ─── Main ────────────────────────────────────────────────────
def main():
    results: dict[str, str] = {}

    # ── Admin ──
    admin_token = login("testadmin_x1", "Admin@123")
    if not admin_token:
        print("[ERROR] Khong login duoc admin"); sys.exit(1)
    A = {"Authorization": f"Bearer {admin_token}"}

    # ── Doctor (non-admin) ──
    doctor_token = login("nguyenvand", "Doctor@123")
    D = {"Authorization": f"Bearer {doctor_token}"} if doctor_token else None

    # ── Ke toan: tim tu salary/staff va doi chieu auth/users ──
    Acc = None
    acct_user = None
    try:
        staff_r  = requests.get(f"{BASE}/api/salary/staff", headers=A)
        eligible = staff_r.json() if staff_r.status_code == 200 else []
        acct_ids = {s["id"] for s in eligible if s.get("role") == "ACCOUNTANT"}
        recept_ids = {s["id"] for s in eligible if s.get("role") == "RECEPTIONIST"}

        users_r  = requests.get(f"{BASE}/api/auth/users", headers=A)
        all_users = users_r.json() if isinstance(users_r.json(), list) \
                    else users_r.json().get("users", [])

        for u in all_users:
            if u.get("id") in acct_ids and u.get("isActive"):
                acct_user = u
                break
            roles = u.get("roles", [])
            role_names = [r.get("name","") if isinstance(r, dict) else r for r in roles]
            if "ACCOUNTANT" in role_names and u.get("isActive"):
                acct_user = u
                break

        if acct_user:
            uid = acct_user["id"]
            rr  = requests.post(f"{BASE}/api/auth/users/{uid}/reset-password",
                                json={"newPassword": "Test@12345"}, headers=A)
            if rr.status_code == 200:
                at = login(acct_user.get("username"), "Test@12345")
                if at:
                    Acc = {"Authorization": f"Bearer {at}"}
                    print(f"[SETUP] Ke toan OK: {acct_user.get('username')}")
    except Exception as e:
        print(f"[SETUP] Loi tim ke toan: {e}")

    # ── Staff IDs ──
    staff_r  = requests.get(f"{BASE}/api/salary/staff", headers=A)
    eligible = staff_r.json() if staff_r.status_code == 200 else []
    receptionist_id = next((s["id"] for s in eligible if s.get("role") == "RECEPTIONIST"), None)
    accountant_id   = next((s["id"] for s in eligible if s.get("role") == "ACCOUNTANT"),   None)
    doctor_id       = 8   # nguyenvand

    print(f"[SETUP] receptionist_id={receptionist_id}  accountant_id={accountant_id}")

    # ── Clean test data truoc khi chay ──
    clean_test_data()

    # ═══════════════════════════════════════════════════════════
    # uc4.1_01 – Admin thiet lap don gia gio thanh cong
    # Dung 2025-01-01 ~ 2025-06-30: khong trung voi data thuc (2026+)
    # ═══════════════════════════════════════════════════════════
    r = requests.post(f"{BASE}/api/salary/hourly-rates",
                      json={"amount": 200_000,
                            "startDate": HR_START1,
                            "endDate":   HR_END1},
                      headers=A)
    ok = r.status_code in (200, 201) and r.json().get("ok")
    results["uc4.1_01"] = "Đạt" if ok else f"Không đạt (HTTP {r.status_code}: {r.text[:150]})"
    print(f"uc4.1_01: {r.status_code} → {results['uc4.1_01']}")

    # ═══════════════════════════════════════════════════════════
    # uc4.1_02 – Admin thiet lap luong co dinh cho le tan
    # ═══════════════════════════════════════════════════════════
    if receptionist_id:
        r = requests.post(f"{BASE}/api/salary/fixed-salaries",
                          json={"userId":    receptionist_id,
                                "amount":    8_000_000,
                                "startDate": FS_START,
                                "endDate":   FS_END},
                          headers=A)
        ok = r.status_code in (200, 201) and r.json().get("ok")
        results["uc4.1_02"] = "Đạt" if ok else f"Không đạt (HTTP {r.status_code}: {r.text[:150]})"
    else:
        results["uc4.1_02"] = "Chưa test (không có lễ tân trong hệ thống)"
    print(f"uc4.1_02: → {results['uc4.1_02']}")

    # ═══════════════════════════════════════════════════════════
    # uc4.1_03 – Admin thiet lap luong co dinh cho ke toan
    # Ke toan (userId=2) co salary tu 2026-01-01; dung 2025 -> khong trung
    # ═══════════════════════════════════════════════════════════
    if accountant_id:
        r = requests.post(f"{BASE}/api/salary/fixed-salaries",
                          json={"userId":    accountant_id,
                                "amount":    10_000_000,
                                "startDate": FS_START,
                                "endDate":   FS_END},
                          headers=A)
        ok = r.status_code in (200, 201) and r.json().get("ok")
        results["uc4.1_03"] = "Đạt" if ok else f"Không đạt (HTTP {r.status_code}: {r.text[:150]})"
    else:
        results["uc4.1_03"] = "Chưa test (không có kế toán trong hệ thống)"
    print(f"uc4.1_03: → {results['uc4.1_03']}")

    # ═══════════════════════════════════════════════════════════
    # uc4.1_04 – That bai khi amount <= 0
    # ═══════════════════════════════════════════════════════════
    r0 = requests.post(f"{BASE}/api/salary/hourly-rates",
                       json={"amount": 0, "startDate": "2025-07-01"}, headers=A)
    r_neg = requests.post(f"{BASE}/api/salary/hourly-rates",
                          json={"amount": -500, "startDate": "2025-07-01"}, headers=A)
    ok = r0.status_code == 400 and r_neg.status_code == 400
    results["uc4.1_04"] = ("Đạt" if ok
                            else f"Không đạt (amount=0→{r0.status_code}, amount=-500→{r_neg.status_code})")
    print(f"uc4.1_04: amount=0→{r0.status_code}, amount=-500→{r_neg.status_code} → {results['uc4.1_04']}")

    # ═══════════════════════════════════════════════════════════
    # uc4.1_05 – That bai khi luong co dinh < 0
    # ═══════════════════════════════════════════════════════════
    if receptionist_id:
        r = requests.post(f"{BASE}/api/salary/fixed-salaries",
                          json={"userId":    receptionist_id,
                                "amount":    -1_000_000,
                                "startDate": "2025-07-01"},
                          headers=A)
        ok = r.status_code == 400
        results["uc4.1_05"] = "Đạt" if ok else f"Không đạt (HTTP {r.status_code}: {r.text[:150]})"
    else:
        results["uc4.1_05"] = "Chưa test (không có lễ tân)"
    print(f"uc4.1_05: → {results['uc4.1_05']}")

    # ═══════════════════════════════════════════════════════════
    # uc4.1_06 – That bai khi khoang thoi gian chong lan
    # uc4.1_01 da tao: 2025-01-01 ~ 2025-06-30
    # Thu tao: 2025-03-01 ~ 2025-09-30 → trung voi record tren → 409
    # ═══════════════════════════════════════════════════════════
    r = requests.post(f"{BASE}/api/salary/hourly-rates",
                      json={"amount": 220_000,
                            "startDate": "2025-03-01",
                            "endDate":   "2025-09-30"},
                      headers=A)
    ok = r.status_code == 409
    results["uc4.1_06"] = "Đạt" if ok else f"Không đạt (HTTP {r.status_code}: {r.text[:150]})"
    print(f"uc4.1_06: {r.status_code} → {results['uc4.1_06']}")

    # ═══════════════════════════════════════════════════════════
    # uc4.1_07 – That bai khi ngay ket thuc <= ngay bat dau
    # ═══════════════════════════════════════════════════════════
    r_eq = requests.post(f"{BASE}/api/salary/hourly-rates",
                         json={"amount": 100_000,
                               "startDate": "2024-06-10",
                               "endDate":   "2024-06-10"},   # bang nhau
                         headers=A)
    r_lt = requests.post(f"{BASE}/api/salary/hourly-rates",
                         json={"amount": 100_000,
                               "startDate": "2024-06-10",
                               "endDate":   "2024-01-01"},   # nho hon
                         headers=A)
    ok = r_eq.status_code == 400 and r_lt.status_code == 400
    results["uc4.1_07"] = ("Đạt" if ok
                            else f"Không đạt (equal→{r_eq.status_code}, less→{r_lt.status_code})")
    print(f"uc4.1_07: equal→{r_eq.status_code}, less→{r_lt.status_code} → {results['uc4.1_07']}")

    # ═══════════════════════════════════════════════════════════
    # uc4.1_08 – Khong cho thiet lap luong co dinh cho bac si
    # ═══════════════════════════════════════════════════════════
    r = requests.post(f"{BASE}/api/salary/fixed-salaries",
                      json={"userId":    doctor_id,
                            "amount":    15_000_000,
                            "startDate": "2025-07-01"},
                      headers=A)
    ok = r.status_code == 400
    results["uc4.1_08"] = "Đạt" if ok else f"Không đạt (HTTP {r.status_code}: {r.text[:150]})"
    print(f"uc4.1_08: {r.status_code} → {results['uc4.1_08']}")

    # ═══════════════════════════════════════════════════════════
    # uc4.1_09 – Ke toan chi xem duoc, khong sua duoc cau hinh
    # GET /api/salary/hourly-rates       → 200 (readOnly=ADMIN|ACCOUNTANT)
    # GET /api/salary/fixed-salaries     → 200
    # POST /api/salary/hourly-rates      → 403 (admin only)
    # POST /api/salary/fixed-salaries    → 403 (admin only)
    # ═══════════════════════════════════════════════════════════
    if Acc:
        r_g1  = requests.get(f"{BASE}/api/salary/hourly-rates",   headers=Acc)
        r_g2  = requests.get(f"{BASE}/api/salary/fixed-salaries",  headers=Acc)
        r_p1  = requests.post(f"{BASE}/api/salary/hourly-rates",
                              json={"amount": 999, "startDate": "2025-08-01"}, headers=Acc)
        r_p2  = requests.post(f"{BASE}/api/salary/fixed-salaries",
                              json={"userId": receptionist_id or 1, "amount": 999,
                                    "startDate": "2025-08-01"}, headers=Acc)
        ok = (r_g1.status_code == 200 and r_g2.status_code == 200
              and r_p1.status_code == 403 and r_p2.status_code == 403)
        results["uc4.1_09"] = ("Đạt" if ok
                                else f"Không đạt (GET hr={r_g1.status_code}, "
                                     f"GET fs={r_g2.status_code}, "
                                     f"POST hr={r_p1.status_code}, "
                                     f"POST fs={r_p2.status_code})")
    else:
        results["uc4.1_09"] = "Chưa test (không có tài khoản kế toán)"
    print(f"uc4.1_09: → {results['uc4.1_09']}")

    # ═══════════════════════════════════════════════════════════
    # uc4.1_10 – Nguoi khong phai Admin khong duoc thiet lap cau hinh luong
    # Dung token bac si / le tan → POST → 403
    # ═══════════════════════════════════════════════════════════
    non_admin = D  # bac si
    if not non_admin and Acc:
        non_admin = Acc  # ke toan cung khong duoc
    if non_admin:
        r = requests.post(f"{BASE}/api/salary/hourly-rates",
                          json={"amount": 200_000, "startDate": "2025-08-01"},
                          headers=non_admin)
        ok = r.status_code == 403
        results["uc4.1_10"] = "Đạt" if ok else f"Không đạt (HTTP {r.status_code}: {r.text[:150]})"
        print(f"uc4.1_10: {r.status_code} → {results['uc4.1_10']}")
    else:
        results["uc4.1_10"] = "Chưa test (không có tài khoản non-admin)"
        print(f"uc4.1_10: → {results['uc4.1_10']}")

    # ═══════════════════════════════════════════════════════════
    # uc4.1_11 – Xem lich su cau hinh luong thanh cong
    # GET /api/salary/hourly-rates   → 200, list
    # GET /api/salary/fixed-salaries → 200, list
    # ═══════════════════════════════════════════════════════════
    r_hr = requests.get(f"{BASE}/api/salary/hourly-rates",   headers=A)
    r_fs = requests.get(f"{BASE}/api/salary/fixed-salaries",  headers=A)
    ok = r_hr.status_code == 200 and r_fs.status_code == 200
    hr_list = r_hr.json() if ok else []
    fs_list = r_fs.json() if ok else []
    results["uc4.1_11"] = ("Đạt" if ok
                            else f"Không đạt (HR={r_hr.status_code}, FS={r_fs.status_code})")
    print(f"uc4.1_11: HR={r_hr.status_code}({len(hr_list)} records), "
          f"FS={r_fs.status_code}({len(fs_list)} records) → {results['uc4.1_11']}")

    # ═══════════════════════════════════════════════════════════
    # uc4.1_12 – Lich su khong bi xoa khi them cau hinh moi
    # Tao them record thu 2: 2025-07-01 ~ 2025-12-31 (khong chong voi 2025-01-01~06-30)
    # Sau do GET pha hai: phai co >= 2 records nam 2025 (test data)
    # va tong >= 3 records (bao gom data thuc 2026)
    # ═══════════════════════════════════════════════════════════
    r2 = requests.post(f"{BASE}/api/salary/hourly-rates",
                       json={"amount":    220_000,
                             "startDate": HR_START2,
                             "endDate":   HR_END2},
                       headers=A)
    r_list = requests.get(f"{BASE}/api/salary/hourly-rates", headers=A)
    if r_list.status_code == 200:
        all_rates   = r_list.json()
        cnt_2025    = sum(1 for x in all_rates if "2025" in str(x.get("startDate", "")))
        total_cnt   = len(all_rates)
        # Lich su duoc giu: phai co ca record 2025-01-01 VA 2025-07-01
        ok = cnt_2025 >= 2 and total_cnt >= 3
        results["uc4.1_12"] = ("Đạt" if ok
                                else f"Không đạt (2025={cnt_2025}, total={total_cnt})")
    else:
        results["uc4.1_12"] = f"Không đạt (GET={r_list.status_code})"
    print(f"uc4.1_12: → {results['uc4.1_12']}")

    # ═══════════════════════════════════════════════════════════
    # uc4.1_13 – Loi CSDL → khong the mo phong qua API
    # ═══════════════════════════════════════════════════════════
    results["uc4.1_13"] = "Chưa test"
    print("uc4.1_13: Không thể mô phỏng lỗi DB qua REST API")

    # ═══════════════════════════════════════════════════════════
    # uc4.1_14 – Ghi audit log khi cau hinh thay doi
    # ═══════════════════════════════════════════════════════════
    cnt = count_audit("CREATE_HOURLY_RATE")
    ok  = cnt > 0
    results["uc4.1_14"] = ("Đạt" if ok
                            else "Không đạt (không có log CREATE_HOURLY_RATE)")
    print(f"uc4.1_14: audit_count={cnt} → {results['uc4.1_14']}")

    # ═══════════════════════════════════════════════════════════
    # Cap nhat Excel
    # ═══════════════════════════════════════════════════════════
    print("\n=== Cap nhat Excel ===")
    try:
        wb = load_workbook(EXCEL)
        ws = wb[SHEET]
        tc_rows: dict[str, int] = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    v = cell.value.strip().lower()
                    if v in results:
                        tc_rows[v] = cell.row
                        break

        for tc_id, row_num in tc_rows.items():
            val         = results[tc_id]
            status_cell = ws.cell(row=row_num, column=7)
            status_cell.value = val
            if val == "Đạt":
                status_cell.fill = GREEN
            elif "Không đạt" in val:
                status_cell.fill = RED
            else:
                status_cell.fill = PatternFill()  # clear fill

        wb.save(EXCEL)
        print(f"Da luu Excel – cap nhat {len(tc_rows)}/{len(results)} test cases")
    except Exception as e:
        print(f"Loi luu Excel: {e}")

    # ═══════════════════════════════════════════════════════════
    print("\n═══════════════ KET QUA UC4.1 ═══════════════")
    dat  = sum(1 for v in results.values() if v == "Đạt")
    kdat = sum(1 for v in results.values() if "Không đạt" in v)
    chua = sum(1 for v in results.values() if "Chưa test" in v)
    for k, v in results.items():
        icon = "✓" if v == "Đạt" else ("✗" if "Không đạt" in v else "○")
        print(f"  {icon} {k}: {v}")
    print(f"\nTong: {dat} Dat | {kdat} Khong dat | {chua} Chua test / 14")

if __name__ == "__main__":
    main()
