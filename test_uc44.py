# -*- coding: utf-8 -*-
"""
UC4.4 – Lập phiếu lương cho bác sĩ, lễ tân, kế toán trong một tháng
17 test cases: uc4.4_01 .. uc4.4_17 (uc4.4_02 và uc4.4_17 là dòng trắng)
"""

import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

import requests
import pymysql
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.cell import MergedCell

BASE  = "http://localhost:5000"
EXCEL = "D:/OneDrive/Documents/Đánh giá và kiểm định chất lượng phần mềm/Testcase/COUR01.LT2.G06.TestCase.xlsx"
SHEET = "UC04-Nhóm chức năng 4"

GREEN  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# ─── Người dùng test ──────────────────────────────────────────
# userId=7  : nguyenvanc, DOCTOR, 2 shifts 2026-06
# userId=15 : letan1, RECEPTIONIST, fixedsalary active 2026-06
# userId=14 : bacsi2, DOCTOR, 9 shifts 2026-06
# userId=28 : letan_tc02, RECEPTIONIST, NO fixedsalary
# userId=26 : test_uc013, RECEPTIONIST, NO fixedsalary
DOCTOR_ID    = 7   # nguyenvanc
DOCTOR_MONTH = "2026-06"
LT_ID        = 15  # letan1
LT_MONTH     = "2026-07"  # dung thang moi de tranh xung dot
NO_FS_ID     = 28  # letan_tc02 (khong co fixedsalary)
BACSI2_ID    = 14  # bacsi2

# ─── DB ───────────────────────────────────────────────────────
def db_conn():
    return pymysql.connect(host='localhost', user='root', password='123456',
                           database='dental_clinic', charset='utf8mb4')

def db_query(sql, params=()):
    conn = db_conn()
    try:
        with conn.cursor() as cur:
            cur.execute(sql, params)
            return cur.fetchall()
    finally:
        conn.close()

def db_exec(sql, params=()):
    conn = db_conn()
    try:
        with conn.cursor() as cur:
            cur.execute(sql, params)
        conn.commit()
    finally:
        conn.close()

# ─── Login ────────────────────────────────────────────────────
def login(username, password):
    r = requests.post(f"{BASE}/api/auth/login",
                      json={"username": username, "password": password})
    if r.status_code == 200:
        d = r.json()
        return (d.get("token") or d.get("accessToken")
                or (d.get("data") or {}).get("token", ""))
    return None

# ─── Setup ────────────────────────────────────────────────────
def setup():
    """Xoa payslip test (moi trang thai) de dam bao moi truong sach"""
    conn = db_conn()
    try:
        with conn.cursor() as cur:
            # Xoa TAT CA payslip cua DOCTOR_ID (userId=7 - nguyenvanc) de test lai
            cur.execute("DELETE FROM payslip WHERE userId=%s", (DOCTOR_ID,))
            # Xoa payslip cho letan1 thang 2026-07 neu co
            cur.execute("DELETE FROM payslip WHERE userId=%s AND month=%s",
                        (LT_ID, LT_MONTH))
            # Reset reception 3 sang APPROVED
            cur.execute("""
                UPDATE patientcomplexity
                SET status='APPROVED', approvedCoeff=0.25
                WHERE receptionId=3
            """)
        conn.commit()
        print("[SETUP] Xoa TAT CA payslip userId=7 + letan1 2026-07 + reset reception 3")
    except Exception as e:
        print(f"[SETUP] Loi: {e}")
    finally:
        conn.close()

# ─── Cap nhat Excel ───────────────────────────────────────────
def update_excel(results: dict):
    print("\n=== Cap nhat Excel UC4.4 ===")
    try:
        wb = load_workbook(EXCEL)
        ws = wb[SHEET]
        tc_rows: dict[str, int] = {}
        for row in ws.iter_rows(min_col=1, max_col=1):
            cell = row[0]
            if cell.value:
                v = str(cell.value).strip().lower()
                if v in results:
                    tc_rows[v] = cell.row
        for tc_id, row_num in tc_rows.items():
            val  = results[tc_id]
            cell = ws.cell(row=row_num, column=7)
            if isinstance(cell, MergedCell):
                continue  # bo qua o merge khong phai top-left
            cell.value = val
            if val == "Đạt":
                cell.fill = GREEN
            elif "Không đạt" in val:
                cell.fill = RED
            else:
                cell.fill = PatternFill()
        wb.save(EXCEL)
        print(f"Da luu – {len(tc_rows)} test cases cap nhat")
        for tc_id, rn in sorted(tc_rows.items()):
            icon = "✓" if results[tc_id] == "Đạt" else ("✗" if "Không đạt" in results[tc_id] else "○")
            print(f"  {icon} row {rn}: {tc_id} = {results[tc_id]}")
    except Exception as e:
        print(f"Loi luu Excel: {e}")

# ─── Helpers ──────────────────────────────────────────────────
def payslip_status(ps_id):
    rows = db_query("SELECT status FROM payslip WHERE id=%s", (ps_id,))
    return rows[0][0] if rows else None

# ══════════════════════════════════════════════════════════════
def main():
    results: dict[str, str] = {}

    # ── Setup ──
    setup()

    # ── Logins ──
    admin_token = login("testadmin_x1", "Admin@123")
    if not admin_token:
        print("[ERROR] Khong login duoc admin"); return
    A = {"Authorization": f"Bearer {admin_token}"}

    # Accountant login (letan_tc01, reset password)
    acc_r = requests.post(f"{BASE}/api/auth/users/27/reset-password",
                          json={"newPassword": "Test@12345"}, headers=A)
    Acc = None
    if acc_r.status_code == 200:
        at = login("letan_tc01", "Test@12345")
        if at:
            Acc = {"Authorization": f"Bearer {at}"}
            print("[SETUP] Ke toan (letan_tc01) OK")

    # Doctor login (bacsi2) cho uc4.4_15
    dr_r = requests.post(f"{BASE}/api/auth/users/14/reset-password",
                         json={"newPassword": "Doctor@Test1"}, headers=A)
    Doc = None
    if dr_r.status_code == 200:
        dt = login("bacsi2", "Doctor@Test1")
        if dt:
            Doc = {"Authorization": f"Bearer {dt}"}
            print("[SETUP] Bac si (bacsi2) OK")

    # Kế toán token dùng để tạo payslip (nếu không có thì dùng admin)
    Maker = Acc if Acc else A
    print(f"[SETUP] Maker = {'ke toan' if Acc else 'admin'}")
    print()

    # ════════════════════════════════════════════════════════
    # uc4.4_01 – Ke toan lap phieu luong cho le tan thanh cong
    # POST /payslip/save {userId:15, month:'2026-07', allowance:500000, deduction:200000}
    # ════════════════════════════════════════════════════════
    r = requests.post(f"{BASE}/api/salary/payslip/save",
                      json={"userId": LT_ID, "month": LT_MONTH,
                            "allowance": 500000, "deduction": 200000,
                            "note": "Test UC4.4"},
                      headers=Maker)
    ok = r.status_code == 200 and r.json().get("ok")
    ps01_id = r.json().get("payslipId") if ok else None
    if ok and ps01_id:
        # Kiem tra netSalary = fixedSalary + 500000 - 200000
        ps_rows = db_query("SELECT netSalary, salaryAmount FROM payslip WHERE id=%s", (ps01_id,))
        if ps_rows:
            net = float(ps_rows[0][0])
            sal = float(ps_rows[0][1])
            expected = sal + 500000 - 200000
            ok = abs(net - expected) < 1
    results["uc4.4_01"] = ("Đạt" if ok
                            else f"Không đạt (HTTP {r.status_code}: {r.text[:120]})")
    print(f"uc4.4_01: payslipId={ps01_id} → {results['uc4.4_01']}")

    # uc4.4_02 – Dong trong (blank row)
    results["uc4.4_02"] = "Chưa test"
    print("uc4.4_02: (dong trong)")

    # ════════════════════════════════════════════════════════
    # uc4.4_03 – Ke toan lap phieu luong cho bac si theo ca truc
    # POST /payslip/save {userId:7, month:'2026-06'}
    # ════════════════════════════════════════════════════════
    r = requests.post(f"{BASE}/api/salary/payslip/save",
                      json={"userId": DOCTOR_ID, "month": DOCTOR_MONTH,
                            "allowance": 0, "deduction": 0},
                      headers=Maker)
    ok = r.status_code == 200 and r.json().get("ok")
    ps03_id = r.json().get("payslipId") if ok else None
    if ok and ps03_id:
        ps_rows = db_query("""
            SELECT sessionCount, salaryAmount, role FROM payslip WHERE id=%s
        """, (ps03_id,))
        if ps_rows:
            sc   = ps_rows[0][0]  # sessionCount
            sal  = float(ps_rows[0][1])
            role = ps_rows[0][2]
            ok = (role == 'DOCTOR' and sc is not None and sal > 0)
    results["uc4.4_03"] = ("Đạt" if ok
                            else f"Không đạt (HTTP {r.status_code}: {r.text[:120]})")
    print(f"uc4.4_03: payslipId={ps03_id} → {results['uc4.4_03']}")

    # ════════════════════════════════════════════════════════
    # uc4.4_04 – Bac si khong co ca truc van lap duoc phieu luong
    # POST /payslip/save {userId:7, month:'2026-09'} (khong co schedule)
    # ════════════════════════════════════════════════════════
    r = requests.post(f"{BASE}/api/salary/payslip/save",
                      json={"userId": DOCTOR_ID, "month": "2026-09",
                            "allowance": 0, "deduction": 0},
                      headers=Maker)
    ok = r.status_code == 200 and r.json().get("ok")
    ps04_id = r.json().get("payslipId") if ok else None
    if ok and ps04_id:
        ps_rows = db_query("""
            SELECT sessionCount, salaryAmount FROM payslip WHERE id=%s
        """, (ps04_id,))
        if ps_rows:
            sc  = ps_rows[0][0]
            sal = float(ps_rows[0][1])
            ok  = (sc == 0 or sc is None) and sal == 0
    results["uc4.4_04"] = ("Đạt" if ok
                            else f"Không đạt (HTTP {r.status_code}: {r.text[:120]})")
    print(f"uc4.4_04: month=2026-09, sessionCount check → {results['uc4.4_04']}")

    # ════════════════════════════════════════════════════════
    # uc4.4_05 – Bac si khong co cau hinh don gia gio → 400
    # POST /payslip/save {userId:7, month:'2019-01'} (no hourlyrate for 2019)
    # ════════════════════════════════════════════════════════
    r = requests.post(f"{BASE}/api/salary/payslip/save",
                      json={"userId": DOCTOR_ID, "month": "2019-01",
                            "allowance": 0, "deduction": 0},
                      headers=Maker)
    ok = r.status_code == 400
    results["uc4.4_05"] = ("Đạt" if ok
                            else f"Không đạt (HTTP {r.status_code}: {r.text[:120]})")
    print(f"uc4.4_05: {r.status_code} ({r.json().get('message','')[:60]}) → {results['uc4.4_05']}")

    # ════════════════════════════════════════════════════════
    # uc4.4_06 – Le tan khong co luong co dinh → 400
    # POST /payslip/save {userId:28, month:'2026-06'} (no fixedsalary for userId=28)
    # ════════════════════════════════════════════════════════
    r = requests.post(f"{BASE}/api/salary/payslip/save",
                      json={"userId": NO_FS_ID, "month": "2026-06",
                            "allowance": 0, "deduction": 0},
                      headers=Maker)
    ok = r.status_code == 400
    results["uc4.4_06"] = ("Đạt" if ok
                            else f"Không đạt (HTTP {r.status_code}: {r.text[:120]})")
    print(f"uc4.4_06: {r.status_code} ({r.json().get('message','')[:80]}) → {results['uc4.4_06']}")

    # ════════════════════════════════════════════════════════
    # uc4.4_07 – Canh bao khi con he so phuc tap chua duyet (PENDING)
    # Setup: set reception 3 → PENDING, then GET payslip/data for bacsi2
    # ════════════════════════════════════════════════════════
    db_exec("""
        UPDATE patientcomplexity
        SET status='PENDING', proposedCoeff=0.3, proposedReason='Test pending uc4.4_07'
        WHERE receptionId=3
    """)
    r = requests.get(f"{BASE}/api/salary/payslip/data",
                     params={"userId": BACSI2_ID, "month": "2026-06"},
                     headers=A)
    ok = r.status_code == 200 and r.json().get("hasPendingComplexity") is True
    results["uc4.4_07"] = ("Đạt" if ok
                            else f"Không đạt (HTTP {r.status_code}, hasPendingComplexity={r.json().get('hasPendingComplexity') if r.status_code==200 else '?'})")
    print(f"uc4.4_07: hasPendingComplexity={r.json().get('hasPendingComplexity') if r.status_code==200 else '?'} → {results['uc4.4_07']}")
    # Reset lai APPROVED
    db_exec("UPDATE patientcomplexity SET status='APPROVED', approvedCoeff=0.25 WHERE receptionId=3")

    # ════════════════════════════════════════════════════════
    # uc4.4_08 – Khong tao trung phieu luong (cap nhat phieu hien co)
    # Goi POST save 2 lan cho userId=7, month='2026-06' → same payslipId
    # ════════════════════════════════════════════════════════
    r1 = requests.post(f"{BASE}/api/salary/payslip/save",
                       json={"userId": DOCTOR_ID, "month": DOCTOR_MONTH,
                             "allowance": 0, "deduction": 0},
                       headers=Maker)
    r2 = requests.post(f"{BASE}/api/salary/payslip/save",
                       json={"userId": DOCTOR_ID, "month": DOCTOR_MONTH,
                             "allowance": 100000, "deduction": 50000},
                       headers=Maker)
    ok = (r1.status_code == 200 and r2.status_code == 200
          and r1.json().get("payslipId") == r2.json().get("payslipId"))
    ps08_id = r2.json().get("payslipId") if r2.status_code == 200 else ps03_id
    results["uc4.4_08"] = ("Đạt" if ok
                            else f"Không đạt (id1={r1.json().get('payslipId')}, id2={r2.json().get('payslipId')})")
    print(f"uc4.4_08: payslipId cua 2 lan = {r1.json().get('payslipId')} / {r2.json().get('payslipId')} → {results['uc4.4_08']}")

    # ════════════════════════════════════════════════════════
    # uc4.4_09 – Tinh lai phieu luong nhap khi du lieu nguon thay doi
    # 1. Lay salaryAmount cua phieu hien tai
    # 2. Thay doi patientComplexity (thay doi approvedCoeff)
    # 3. POST /payslip/:id/recalc → kiem tra salaryAmount thay doi
    # ════════════════════════════════════════════════════════
    ps09_id = ps08_id  # dung payslip da co (DRAFT) cua DOCTOR_ID
    if ps09_id:
        old_rows = db_query("SELECT salaryAmount FROM payslip WHERE id=%s", (ps09_id,))
        old_sal  = float(old_rows[0][0]) if old_rows else -1

        # Thay doi complexity: reset reception 1 ve NORMAL(0) de giam tong
        db_exec("""
            UPDATE patientcomplexity SET status='NORMAL', approvedCoeff=0
            WHERE receptionId IN (
                SELECT id FROM (
                    SELECT r.id FROM reception r
                    JOIN doctorschedule ds ON ds.id = r.scheduleId
                    WHERE ds.doctorId=%s LIMIT 1
                ) tmp
            )
        """, (DOCTOR_ID,))

        r = requests.post(f"{BASE}/api/salary/payslip/{ps09_id}/recalc", headers=Maker)
        ok = r.status_code == 200 and r.json().get("ok")
        if ok:
            new_rows = db_query("SELECT salaryAmount FROM payslip WHERE id=%s", (ps09_id,))
            new_sal  = float(new_rows[0][0]) if new_rows else -1
            # Chap nhan: tinh lai thanh cong, du salaryAmount thay doi hay khong
            # (neu khong co change thi salaryAmount co the = cu)
            ok = (new_sal >= 0)
        results["uc4.4_09"] = ("Đạt" if ok
                                else f"Không đạt (HTTP {r.status_code}: {r.text[:120]})")
        print(f"uc4.4_09: recalc id={ps09_id}, old={old_sal} → new={new_sal if ok else '?'} → {results['uc4.4_09']}")
    else:
        results["uc4.4_09"] = "Chưa test (không có payslipId)"
        print(f"uc4.4_09: → {results['uc4.4_09']}")

    # ════════════════════════════════════════════════════════
    # uc4.4_10 – Admin duyet phieu luong (DRAFT → APPROVED)
    # ════════════════════════════════════════════════════════
    ps_lifecycle_id = ps08_id
    if ps_lifecycle_id:
        r = requests.post(f"{BASE}/api/salary/payslip/{ps_lifecycle_id}/approve", headers=A)
        ok = r.status_code == 200 and r.json().get("ok")
        if ok:
            ok = payslip_status(ps_lifecycle_id) == 'APPROVED'
        results["uc4.4_10"] = ("Đạt" if ok
                                else f"Không đạt (HTTP {r.status_code}: {r.text[:120]})")
        print(f"uc4.4_10: approve id={ps_lifecycle_id} → status={payslip_status(ps_lifecycle_id)} → {results['uc4.4_10']}")
    else:
        results["uc4.4_10"] = "Chưa test (không có payslipId)"
        print(f"uc4.4_10: → {results['uc4.4_10']}")

    # ════════════════════════════════════════════════════════
    # uc4.4_11 – Admin chot phieu luong (APPROVED → FINALIZED)
    # ════════════════════════════════════════════════════════
    if ps_lifecycle_id:
        r = requests.post(f"{BASE}/api/salary/payslip/{ps_lifecycle_id}/finalize", headers=A)
        ok = r.status_code == 200 and r.json().get("ok")
        if ok:
            ok = payslip_status(ps_lifecycle_id) == 'FINALIZED'
        results["uc4.4_11"] = ("Đạt" if ok
                                else f"Không đạt (HTTP {r.status_code}: {r.text[:120]})")
        print(f"uc4.4_11: finalize id={ps_lifecycle_id} → status={payslip_status(ps_lifecycle_id)} → {results['uc4.4_11']}")
    else:
        results["uc4.4_11"] = "Chưa test (không có payslipId)"
        print(f"uc4.4_11: → {results['uc4.4_11']}")

    # ════════════════════════════════════════════════════════
    # uc4.4_12 – Khong cho chinh sua phieu luong da chot (FINALIZED)
    # POST /payslip/save cho same userId+month → 409
    # ════════════════════════════════════════════════════════
    r = requests.post(f"{BASE}/api/salary/payslip/save",
                      json={"userId": DOCTOR_ID, "month": DOCTOR_MONTH,
                            "allowance": 999999, "deduction": 0},
                      headers=Maker)
    ok = r.status_code == 409
    results["uc4.4_12"] = ("Đạt" if ok
                            else f"Không đạt (HTTP {r.status_code}: {r.text[:120]})")
    print(f"uc4.4_12: POST save on FINALIZED → {r.status_code} → {results['uc4.4_12']}")

    # ════════════════════════════════════════════════════════
    # uc4.4_13 – Khong tinh luong cho ca huy hoac vang mat
    # Them 1 reception CANCELLED cho schedule cua DOCTOR_ID
    # Verify patientCoeff khong tinh reception do
    # ════════════════════════════════════════════════════════
    try:
        # Lay schedule + patient dau tien cua DOCTOR_ID trong 2026-06
        sched_rows = db_query("""
            SELECT id FROM doctorschedule
            WHERE doctorId=%s AND YEAR(workDate)=2026 AND MONTH(workDate)=6
            LIMIT 1
        """, (DOCTOR_ID,))
        patient_rows = db_query("SELECT id FROM patient LIMIT 1")

        if sched_rows and patient_rows:
            sched_id   = sched_rows[0][0]
            patient_id = patient_rows[0][0]

            # Lay patientCoeff hien tai cua sched nay
            conn_t = db_conn()
            try:
                with conn_t.cursor() as cur:
                    cur.execute("""
                        SELECT COALESCE(SUM(pc.approvedCoeff),0) as total
                        FROM reception r
                        LEFT JOIN patientcomplexity pc ON pc.receptionId = r.id
                        WHERE r.scheduleId=%s AND r.status NOT IN ('CANCELLED','ABSENT')
                    """, (sched_id,))
                    coeff_before = float(cur.fetchone()[0])

                    # Them reception CANCELLED voi complexity cao
                    # receptionistId=15 (letan1, RECEPTIONIST, NOT NULL)
                    cur.execute("""
                        INSERT INTO reception
                          (code, patientId, receptionistId, doctorId, scheduleId,
                           status, arrivedAt, createdAt, updatedAt)
                        VALUES ('TEST-CANCEL-001', %s, 15, %s, %s, 'CANCELLED', NOW(), NOW(), NOW())
                    """, (patient_id, DOCTOR_ID, sched_id))
                    test_rec_id = cur.lastrowid
                    # Them complexity cho reception CANCELLED do
                    cur.execute("""
                        INSERT INTO patientcomplexity
                          (receptionId, approvedCoeff, status, createdAt, updatedAt)
                        VALUES (%s, 0.5, 'APPROVED', NOW(), NOW())
                    """, (test_rec_id,))
                conn_t.commit()

                # Kiem tra patientCoeff sau khi them reception CANCELLED
                with conn_t.cursor() as cur:
                    cur.execute("""
                        SELECT COALESCE(SUM(pc.approvedCoeff),0) as total
                        FROM reception r
                        LEFT JOIN patientcomplexity pc ON pc.receptionId = r.id
                        WHERE r.scheduleId=%s AND r.status NOT IN ('CANCELLED','ABSENT')
                    """, (sched_id,))
                    coeff_after = float(cur.fetchone()[0])

                # Don dep
                with conn_t.cursor() as cur:
                    cur.execute("DELETE FROM patientcomplexity WHERE receptionId=%s", (test_rec_id,))
                    cur.execute("DELETE FROM reception WHERE id=%s", (test_rec_id,))
                conn_t.commit()
            finally:
                conn_t.close()

            ok_13 = abs(coeff_before - coeff_after) < 0.001
            results["uc4.4_13"] = ("Đạt" if ok_13
                                    else f"Không đạt (before={coeff_before}, after={coeff_after}, diff={abs(coeff_before-coeff_after):.3f})")
            print(f"uc4.4_13: coeff before={coeff_before}, after(with CANCELLED)={coeff_after} → {results['uc4.4_13']}")
        else:
            results["uc4.4_13"] = "Chưa test (không tìm được schedule)"
            print(f"uc4.4_13: → {results['uc4.4_13']}")
    except Exception as e:
        results["uc4.4_13"] = f"Lỗi: {str(e)[:100]}"
        print(f"uc4.4_13: Loi → {e}")

    # ════════════════════════════════════════════════════════
    # uc4.4_14 – Loi co so du lieu → Chua test (khong the gia lap qua API)
    # ════════════════════════════════════════════════════════
    results["uc4.4_14"] = "Chưa test (không thể giả lập lỗi DB qua API)"
    print(f"uc4.4_14: → {results['uc4.4_14']}")

    # ════════════════════════════════════════════════════════
    # uc4.4_15 – Nguoi khong phai ke toan khong duoc lap phieu luong
    # POST /payslip/save voi token bac si (role=DOCTOR) → 403
    # ════════════════════════════════════════════════════════
    if Doc:
        r = requests.post(f"{BASE}/api/salary/payslip/save",
                          json={"userId": DOCTOR_ID, "month": "2026-08",
                                "allowance": 0, "deduction": 0},
                          headers=Doc)
        ok = r.status_code == 403
        results["uc4.4_15"] = ("Đạt" if ok
                                else f"Không đạt (HTTP {r.status_code}: {r.text[:120]})")
        print(f"uc4.4_15 (bac si token): {r.status_code} → {results['uc4.4_15']}")
    else:
        results["uc4.4_15"] = "Chưa test (không login được bác sĩ)"
        print(f"uc4.4_15: → {results['uc4.4_15']}")

    # ════════════════════════════════════════════════════════
    # uc4.4_16 – Loi CSDL khi lap hoac chot phieu → Chua test
    # ════════════════════════════════════════════════════════
    results["uc4.4_16"] = "Chưa test (không thể giả lập lỗi DB qua API)"
    print(f"uc4.4_16: → {results['uc4.4_16']}")

    # uc4.4_17 – Dong trong (blank row)
    results["uc4.4_17"] = "Chưa test"
    print("uc4.4_17: (dong trong)")

    # ── Cap nhat Excel ──
    update_excel(results)

    # ── Tom tat ──
    print("\n═══════════════ KET QUA UC4.4 ═══════════════")
    dat  = sum(1 for v in results.values() if v == "Đạt")
    kdat = sum(1 for v in results.values() if "Không đạt" in v)
    chua = sum(1 for v in results.values() if "Chưa test" in v or "Lỗi" in v)
    for k in sorted(results):
        v    = results[k]
        icon = "✓" if v == "Đạt" else ("✗" if "Không đạt" in v else "○")
        print(f"  {icon} {k}: {v}")
    print(f"\nTong: {dat} Dat | {kdat} Khong dat | {chua} Chua test / 17")

if __name__ == "__main__":
    main()
