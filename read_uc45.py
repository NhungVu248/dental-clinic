import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
from openpyxl import load_workbook

EXCEL = "D:/OneDrive/Documents/Đánh giá và kiểm định chất lượng phần mềm/Testcase/COUR01.LT2.G06.TestCase.xlsx"
SHEET = "UC04-Nhóm chức năng 4"
wb = load_workbook(EXCEL)
ws = wb[SHEET]
print(f"Max row: {ws.max_row}\n")

in_uc45 = False
for rn in range(1, ws.max_row + 1):
    a = ws.cell(row=rn, column=1).value
    if a and "uc4.5" in str(a).lower():
        in_uc45 = True
    if in_uc45:
        b = ws.cell(row=rn, column=2).value
        c = ws.cell(row=rn, column=3).value
        d = ws.cell(row=rn, column=4).value
        e = ws.cell(row=rn, column=5).value
        f = ws.cell(row=rn, column=6).value
        g = ws.cell(row=rn, column=7).value
        print(f"Row {rn}: A={str(a)[:70]}")
        if b: print(f"  B={str(b)[:100]}")
        if c: print(f"  C={str(c)[:120]}")
        if d: print(f"  D={str(d)[:120]}")
        if e: print(f"  E={str(e)[:100]}")
        if f: print(f"  F={str(f)[:120]}")
        print(f"  G={g}")
        print()
    if in_uc45 and a and str(a).strip().startswith("UC4.6"):
        break
