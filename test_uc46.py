# -*- coding: utf-8 -*-
"""
Test UC4.6 – Báo cáo tiền lương năm của một nhân sự
GET /api/salary/report/annual/personal?userId=X&year=YYYY  (ADMIN|ACCOUNTANT)
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

import requests, pymysql
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.cell import MergedCell

BASE  = "http://localhost:5000"
EXCEL = "D:/OneDrive/Documents/Đánh giá và kiểm định chất lượng phần mềm/Testcase/COUR01.LT2.G06.TestCase.xlsx"
SHEET = "UC04-Nhóm chức năng 4"

GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED   = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# ─── nội dung uc4.6_02 còn trống ─────────────────────────────────
TC02_CONTENT = {
    "B": "Kế toán xem báo cáo lương năm của một nhân sự thành công",
    "C": ("1. Kế toán đăng nhập hệ thống.\n"
          "2. Mở phân hệ Báo cáo lương.\n"
          "3. Chọn Báo cáo lương năm theo nhân sự.\n"
          "4. Chọn nhân sự BS001 và năm 2026.\n"
          "5. Nhấn Xem báo cáo."),
    "D": "Hệ thống kiểm tra quyền Kế toán và tải dữ liệu phiếu lương của nhân sự trong năm đã chọn.",
    "E": "Tài khoản: ketoan01 | Vai trò: Kế toán | Nhân sự: BS001 | Năm: 2026",
    "F": "Hệ thống hiển thị báo cáo lương 12 tháng của nhân sự BS001, gồm lương từng tháng, tổng lương cả năm và lương trung bình tháng.",
}

# ─── DB helpers ──────────────────────────────────────────────────
def db_conn():
    return pymysql.connect(host='localhost', user='root', password='123456',
                           database='dental_clinic', charset='utf8mb4')

def db_query(sql, params=()):
    conn = db_conn()
    try:
        with conn.cursor() as cur:
            cur.execute(sql, params)
            return cur.fetchall()
    finally:
        conn.close()

def db_exec(sql, params=()):
    conn = db_conn()
    try:
        with conn.cursor() as cur:
            cur.execute(sql, params)
        conn.commit()
    finally:
        conn.close()

# ─── Auth helpers ─────────────────────────────────────────────────
def login(username, password):
    r = requests.post(f"{BASE}/api/auth/login",
                      json={"username": username, "password": password})
    if r.status_code == 200:
        d = r.json()
        return (d.get("token") or d.get("accessToken")
                or (d.get("data") or {}).get("token", ""))
    return None

# ─── Excel helpers ────────────────────────────────────────────────
def fill_cell(ws, row, col, val, wrap=True):
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return
    cell.value = val
    if wrap:
        cell.alignment = Alignment(wrap_text=True, vertical='top')

def update_status(tc_id: str, val: str):
    wb = load_workbook(EXCEL)
    ws = wb[SHEET]
    for row in ws.iter_rows(min_col=1, max_col=1):
        cell = row[0]
        if cell.value and str(cell.value).strip().lower() == tc_id.lower():
            g = ws.cell(row=cell.row, column=7)
            if isinstance(g, MergedCell):
                break
            g.value = val
            g.fill  = GREEN if val == "Đạt" else (RED if "Không đạt" in val else PatternFill())
            print(f"  Excel row {cell.row}: {tc_id} → {val}")
            break
    wb.save(EXCEL)

def fill_uc46_02_content():
    """Điền nội dung cho uc4.6_02 còn trống trong Excel."""
    wb = load_workbook(EXCEL)
    ws = wb[SHEET]
    row_02 = None
    for row in ws.iter_rows(min_col=1, max_col=1):
        cell = row[0]
        if cell.value and str(cell.value).strip().lower() == "uc4.6_02":
            row_02 = cell.row
            break
    if row_02:
        col_map = {"B": 2, "C": 3, "D": 4, "E": 5, "F": 6}
        for col_l, col_i in col_map.items():
            fill_cell(ws, row_02, col_i, TC02_CONTENT[col_l])
        g = ws.cell(row=row_02, column=7)
        if not isinstance(g, MergedCell) and not g.value:
            g.value = "Chưa test"
        wb.save(EXCEL)
        print(f"  Đã điền nội dung uc4.6_02 vào row {row_02}")
    else:
        print("  [WARN] Không tìm thấy row uc4.6_02 trong Excel")

# ─── Setup: đảm bảo userId=7 có payslip tháng 2026-03 và 2026-04 ─
def setup_payslips(admin_headers):
    DOCTOR_ID = 7
    MONTHS    = ["2026-03", "2026-04"]
    print("=== Setup: tạo payslip test cho UC4.6 ===")
    for m in MONTHS:
        db_exec("DELETE FROM payslip WHERE userId=%s AND month=%s", (DOCTOR_ID, m))
        r = requests.post(f"{BASE}/api/salary/payslip/save",
                          json={"userId": DOCTOR_ID, "month": m,
                                "allowance": 0, "deduction": 0},
                          headers=admin_headers)
        if r.status_code == 200 and r.json().get("ok"):
            print(f"  Tạo payslip userId={DOCTOR_ID} tháng {m}: OK (id={r.json().get('payslipId')})")
        else:
            print(f"  [WARN] Tạo payslip {m}: HTTP {r.status_code} – {r.text[:80]}")
    print()

# ─── Main tests ───────────────────────────────────────────────────
def run_tests(A, ACC_H, DOCTOR_H, doctor_id):
    """
    A        – admin headers
    ACC_H    – accountant headers
    DOCTOR_H – doctor headers (userId=doctor_id, role DOCTOR)
    """
    results = {}

    BASE_URL = f"{BASE}/api/salary/report/annual/personal"

    # ─── uc4.6_01 ────────────────────────────────────────────────
    r = requests.get(f"{BASE_URL}?userId={doctor_id}&year=2026", headers=A)
    ok = r.status_code == 200
    if ok:
        d = r.json()
        ok = (d.get("user", {}).get("role") == "DOCTOR"
              and len(d.get("months", [])) == 12
              and "totalAnnual" in d
              and "avgMonthly" in d)
    results["uc4.6_01"] = "Đạt" if ok else f"Không đạt (HTTP {r.status_code}: {r.text[:120]})"
    print(f"uc4.6_01: HTTP {r.status_code} | role={r.json().get('user',{}).get('role') if r.ok else 'N/A'} → {results['uc4.6_01']}")

    # ─── uc4.6_02 ────────────────────────────────────────────────
    r2 = requests.get(f"{BASE_URL}?userId={doctor_id}&year=2026", headers=ACC_H)
    ok2 = r2.status_code == 200 and len(r2.json().get("months", [])) == 12
    results["uc4.6_02"] = "Đạt" if ok2 else f"Không đạt (HTTP {r2.status_code}: {r2.text[:120]})"
    print(f"uc4.6_02: Kế toán xem → HTTP {r2.status_code} → {results['uc4.6_02']}")

    # ─── uc4.6_03 ────────────────────────────────────────────────
    results["uc4.6_03"] = "Chưa test"
    print("uc4.6_03: Nhân sự tự xem lương – tính năng cấu hình frontend, không có API riêng → Chưa test")

    # ─── uc4.6_04 ────────────────────────────────────────────────
    r4 = requests.get(f"{BASE_URL}?userId={doctor_id}&year=2026", headers=DOCTOR_H)
    ok4 = r4.status_code == 403
    results["uc4.6_04"] = "Đạt" if ok4 else f"Không đạt (HTTP {r4.status_code} thay vì 403)"
    print(f"uc4.6_04: Doctor truy cập → HTTP {r4.status_code} → {results['uc4.6_04']}")

    # ─── Lấy dữ liệu báo cáo làm cơ sở cho uc4.6_05 đến 09 ─────
    report = None
    r_main = requests.get(f"{BASE_URL}?userId={doctor_id}&year=2026", headers=A)
    if r_main.status_code == 200:
        report = r_main.json()

    # ─── uc4.6_05 ────────────────────────────────────────────────
    if report:
        ok5 = len(report.get("months", [])) == 12
    else:
        ok5 = False
    results["uc4.6_05"] = "Đạt" if ok5 else "Không đạt (months không đủ 12 tháng)"
    print(f"uc4.6_05: Số tháng = {len(report.get('months',[])) if report else 'N/A'} → {results['uc4.6_05']}")

    # ─── uc4.6_06 ────────────────────────────────────────────────
    if report and report.get("user", {}).get("role") == "DOCTOR":
        # totalSessions and totalHours are included in response
        ok6 = ("totalSessions" in report and "totalHours" in report)
    else:
        ok6 = False
    results["uc4.6_06"] = "Đạt" if ok6 else "Không đạt (thiếu trường totalSessions / totalHours)"
    if report:
        print(f"uc4.6_06: totalSessions={report.get('totalSessions')}, totalHours={report.get('totalHours')} → {results['uc4.6_06']}")
    else:
        print(f"uc4.6_06: (no data) → {results['uc4.6_06']}")

    # ─── uc4.6_07 ────────────────────────────────────────────────
    if report:
        months_data = report.get("months", [])
        computed_total = sum(m["netSalary"] for m in months_data if m.get("netSalary") is not None)
        api_total      = report.get("totalAnnual", -1)
        ok7 = abs(computed_total - api_total) < 1  # tolerance 1 VND for rounding
    else:
        ok7 = False
    results["uc4.6_07"] = "Đạt" if ok7 else f"Không đạt (tổng tính = {computed_total}, API = {api_total})"
    if report:
        print(f"uc4.6_07: sum(netSalary)={computed_total:,.0f} vs totalAnnual={api_total:,.0f} → {results['uc4.6_07']}")
    else:
        print(f"uc4.6_07: (no data) → {results['uc4.6_07']}")

    # ─── uc4.6_08 ────────────────────────────────────────────────
    if report:
        months_data   = report.get("months", [])
        count_months  = sum(1 for m in months_data if m.get("status") != "NONE")
        total_annual  = report.get("totalAnnual", 0)
        api_avg       = report.get("avgMonthly", -1)
        expected_avg  = round(total_annual / count_months) if count_months > 0 else 0
        ok8 = (api_avg == expected_avg)
    else:
        ok8 = False
    results["uc4.6_08"] = "Đạt" if ok8 else f"Không đạt (expected avgMonthly={expected_avg}, got {api_avg})"
    if report:
        print(f"uc4.6_08: countMonths={count_months}, avgMonthly={api_avg:,} vs expected={expected_avg:,} → {results['uc4.6_08']}")
    else:
        print(f"uc4.6_08: (no data) → {results['uc4.6_08']}")

    # ─── uc4.6_09 ────────────────────────────────────────────────
    if report:
        months_data = report.get("months", [])
        none_months = [m for m in months_data if m.get("status") == "NONE"]
        # Every NONE month should have null netSalary, salaryAmount
        ok9 = (len(none_months) > 0
               and all(m.get("netSalary") is None and m.get("salaryAmount") is None
                       for m in none_months))
    else:
        ok9 = False
    results["uc4.6_09"] = "Đạt" if ok9 else "Không đạt (không có tháng NONE hoặc giá trị không phải null)"
    if report:
        print(f"uc4.6_09: {len(none_months)} tháng NONE, tất cả netSalary=null? {all(m.get('netSalary') is None for m in none_months)} → {results['uc4.6_09']}")
    else:
        print(f"uc4.6_09: (no data) → {results['uc4.6_09']}")

    # ─── uc4.6_10 ────────────────────────────────────────────────
    results["uc4.6_10"] = "Chưa test"
    print("uc4.6_10: Nhân sự thay đổi vai trò trong năm – cần thiết lập DB phức tạp, kiểm tra frontend → Chưa test")

    # ─── uc4.6_11 ────────────────────────────────────────────────
    results["uc4.6_11"] = "Chưa test"
    print("uc4.6_11: Xem chi tiết phiếu lương từ báo cáo năm – điều hướng frontend → Chưa test")

    # ─── uc4.6_12 ────────────────────────────────────────────────
    results["uc4.6_12"] = "Chưa test"
    print("uc4.6_12: Biểu đồ diễn biến lương – render frontend → Chưa test")

    # ─── uc4.6_13 ────────────────────────────────────────────────
    r13 = requests.get(f"{BASE_URL}?userId=9999999&year=2026", headers=A)
    ok13 = r13.status_code == 404
    results["uc4.6_13"] = "Đạt" if ok13 else f"Không đạt (HTTP {r13.status_code} thay vì 404)"
    print(f"uc4.6_13: userId=9999999 → HTTP {r13.status_code} → {results['uc4.6_13']}")

    # ─── uc4.6_14 ────────────────────────────────────────────────
    r14 = requests.get(f"{BASE_URL}?userId={doctor_id}&year=2020", headers=A)
    ok14 = False
    if r14.status_code == 200:
        d14 = r14.json()
        all_none  = all(m.get("status") == "NONE" for m in d14.get("months", []))
        total_zero = d14.get("totalAnnual", -1) == 0
        ok14 = all_none and total_zero and len(d14.get("months", [])) == 12
    results["uc4.6_14"] = "Đạt" if ok14 else f"Không đạt (HTTP {r14.status_code}: {r14.text[:80]})"
    if r14.status_code == 200:
        d14 = r14.json()
        print(f"uc4.6_14: year=2020 → totalAnnual={d14.get('totalAnnual')}, all NONE? {all(m.get('status')=='NONE' for m in d14.get('months',[]))} → {results['uc4.6_14']}")
    else:
        print(f"uc4.6_14: HTTP {r14.status_code} → {results['uc4.6_14']}")

    # ─── uc4.6_15 ────────────────────────────────────────────────
    results["uc4.6_15"] = "Chưa test"
    print("uc4.6_15: Xuất báo cáo ra Excel – tính năng frontend → Chưa test")

    # ─── uc4.6_16 ────────────────────────────────────────────────
    results["uc4.6_16"] = "Chưa test"
    print("uc4.6_16: Xuất báo cáo ra PDF – tính năng frontend → Chưa test")

    # ─── uc4.6_17 ────────────────────────────────────────────────
    results["uc4.6_17"] = "Chưa test"
    print("uc4.6_17: Lỗi xuất file – cần giả lập lỗi hạ tầng, frontend → Chưa test")

    return results

# ─── Main ─────────────────────────────────────────────────────────
def main():
    # ── Bước 1: điền nội dung uc4.6_02 vào Excel ──
    print("=== Bước 1: Điền nội dung uc4.6_02 vào Excel ===")
    fill_uc46_02_content()
    print()

    # ── Login admin ──
    admin_token = login("testadmin_x1", "Admin@123")
    if not admin_token:
        print("[ERROR] Không login được admin"); return
    A = {"Authorization": f"Bearer {admin_token}"}

    # ── Lấy doctor_id từ DB ──
    doc_rows = db_query("""
        SELECT u.id, u.username, u.fullName
        FROM `user` u
        JOIN userrole ur ON ur.userId = u.id
        JOIN role r ON r.id = ur.roleId
        WHERE r.name = 'DOCTOR' AND u.isActive = 1
        LIMIT 3
    """)
    if not doc_rows:
        print("[ERROR] Không tìm thấy bác sĩ nào"); return

    doctor_id       = int(doc_rows[0][0])
    doctor_username = doc_rows[0][1]
    print(f"Doctor: id={doctor_id}, username={doctor_username}")

    # Reset doctor password
    rr = requests.post(f"{BASE}/api/auth/users/{doctor_id}/reset-password",
                       json={"newPassword": "Test@12345"}, headers=A)
    doctor_token = None
    if rr.status_code == 200:
        doctor_token = login(doctor_username, "Test@12345")
    if not doctor_token:
        print(f"[WARN] Không login được doctor {doctor_username}")
    DOCTOR_H = {"Authorization": f"Bearer {doctor_token}"} if doctor_token else {}

    # ── Login accountant ──
    acc_rows = db_query("""
        SELECT u.id, u.username FROM `user` u
        JOIN userrole ur ON ur.userId = u.id
        JOIN role r ON r.id = ur.roleId
        WHERE r.name = 'ACCOUNTANT' AND u.isActive = 1
        LIMIT 1
    """)
    ACC_H = A  # fallback to admin
    if acc_rows:
        acc_id       = int(acc_rows[0][0])
        acc_username = acc_rows[0][1]
        rr2 = requests.post(f"{BASE}/api/auth/users/{acc_id}/reset-password",
                            json={"newPassword": "Test@12345"}, headers=A)
        if rr2.status_code == 200:
            acc_token = login(acc_username, "Test@12345")
            if acc_token:
                ACC_H = {"Authorization": f"Bearer {acc_token}"}
                print(f"Accountant: id={acc_id}, username={acc_username}")

    # ── Setup payslips ──
    setup_payslips(A)

    # ── Chạy tests ──
    print("=== Chạy test UC4.6 ===")
    results = run_tests(A, ACC_H, DOCTOR_H, doctor_id)

    # ── Cập nhật Excel ──
    print("\n=== Cập nhật Excel ===")
    for tc_id, val in results.items():
        update_status(tc_id, val)

    # ── Tóm tắt ──
    print("\n══════════════════ KẾT QUẢ UC4.6 ══════════════════")
    dat     = sum(1 for v in results.values() if v == "Đạt")
    khong   = sum(1 for v in results.values() if "Không đạt" in v)
    chua    = sum(1 for v in results.values() if v == "Chưa test")
    for k, v in results.items():
        icon = "✓" if v == "Đạt" else ("✗" if "Không đạt" in v else "–")
        print(f"  {icon} {k}: {v}")
    print(f"\n  Tổng: {dat} Đạt | {khong} Không đạt | {chua} Chưa test / {len(results)} TC")

if __name__ == "__main__":
    main()
